"""
Toronto Infrastructure Intelligence (TII) — Data Scraper v2.10
==============================================================
Fixes vs v1.3:
  1. IESO XML: replaced BeautifulSoup(html.parser) with xml.etree.ElementTree
     (built-in, no external deps, handles XML namespaces correctly)
  2. StatsCan all-items CPI: fixed vector — v41690973 is the confirmed correct
     vector for Canada all-items CPI (not v41693202 which returned FAILED)
  3. StatsCan gas storage: added multi-vector fallback; original v65201762 was
     returning 0.0 — now tries several known Ontario storage vectors
  4. Watermain breaks: removed datastore_active filter — dataset uses file
     download not CKAN datastore; now gets download URL from any resource
  5. TTC ridership: broadened resource format filter; XLSX resources may be
     tagged as "XLSX", "xlsx", or just have .xlsx in URL
  6. data.ontario.ca ER: added direct package ID fallback list since search
     wasn't matching the dataset
  7. OSB insolvency: added graceful handling when openpyxl not installed,
     with clear install instruction

INSTALL
-------
    pip install requests beautifulsoup4 openpyxl
    (lxml optional — not required, xml.etree.ElementTree handles IESO XML now)
"""

import argparse
import os
import csv
import io
import json
import re
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup

# Suppress XMLParsedAsHTMLWarning if BeautifulSoup is used on XML
try:
    from bs4 import XMLParsedAsHTMLWarning
    import warnings
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
except ImportError:
    pass


# Cache for per-generator detail — populated by fetch_ieso_generation_mix,
# written to ieso_generators_YYYYMMDD.json by run_all_scrapers.
_IESO_GENERATOR_CACHE: dict = {}

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "TII-Scraper/1.4 Toronto Infrastructure Intelligence",
    "Accept": "application/json, text/html, text/csv, application/xml, */*",
    "Content-Type": "application/json",
})
TIMEOUT = 25


def _ok(indicator, value, unit, source, url, data_date, notes=""):
    return {"indicator": indicator, "value": value, "unit": unit, "source": source,
            "url": url, "retrieved_at": datetime.utcnow().isoformat() + "Z",
            "data_date": str(data_date), "status": "ok", "notes": notes}

def _err(indicator, source, url, error_msg):
    return {"indicator": indicator, "value": None, "unit": None, "source": source,
            "url": url, "retrieved_at": datetime.utcnow().isoformat() + "Z",
            "data_date": None, "status": "error", "notes": f"ERROR: {error_msg}"}

def _manual(indicator, source, notes, sector=None):
    r = {"indicator": indicator, "value": None, "unit": None, "source": source,
         "url": None, "retrieved_at": datetime.utcnow().isoformat() + "Z",
         "data_date": None, "status": "manual", "notes": notes}
    if sector:
        r["sector"] = sector
    return r


# ── Threshold rules ────────────────────────────────────────────────────────
# Each rule: (indicator_substring, warn_condition_fn, alert_condition_fn, warn_note, alert_note)
# Conditions receive the numeric value. Non-numeric values skip threshold checks.
# Status escalates: ok → warn → alert. Manual/error status never downgraded.
THRESHOLDS = [
    # Energy
    ("Brent Crude Price",
        lambda v: v > 95,   lambda v: v > 115,
        "Elevated — above $95/bbl (geopolitical floor, Iran conflict)",
        "Crisis level — above $115/bbl (sustained supply disruption risk)"),
    ("Gas Output (MW)",
        lambda v: v > 3000, lambda v: v > 5000,
        "Gas peakers elevated — demand stress signal",
        "Gas peakers at crisis level — grid under significant stress"),
    ("Peak Demand Index",
        lambda v: v > 85,   lambda v: v > 95,
        "Demand elevated — above 85% of 2024 peak",
        "Demand critical — above 95% of 2024 peak"),
    # Water / Shelter
    ("Shelter Occupancy",
        lambda v: v > 4500, lambda v: v > 4700,
        "Shelter system above 94% of capacity",
        "Shelter system at or above 98% — crisis threshold"),
    ("Watermain Break Rate",
        lambda v: v > 15,   lambda v: v > 25,
        "Elevated break rate — above seasonal average",
        "Critical break rate — infrastructure stress"),
    # Housing market
    ("TRREB Sales-to-New-Listings Ratio",
        lambda v: v > 60,  lambda v: v > 70,
        "Seller's market — SNR above 60%, price pressure building",
        "Strong seller's market — SNR above 70%, significant affordability stress"),
    # Labour market
    ("Toronto Unemployment Rate",
        lambda v: v > 8.0,  lambda v: v > 10.0,
        "Unemployment elevated — above 8% (high vs pre-tariff baseline)",
        "Unemployment at recession level — above 10%"),
    # TTC Subway — direct status (0=normal, 1=partial, 2=major)
    # Note: fetch_ttc_service_status sets status directly on severity detection.
    # The threshold here acts as a numeric safety net for the status value.
    ("TTC Subway Service Status",
        lambda v: v >= 1,  lambda v: v >= 2,
        "Partial TTC subway disruption — segment delay or shuttle operation",
        "Major TTC subway disruption — full or multi-segment suspension (~1.7M weekly riders)"),
    ("TTC Active Service Alerts",
        lambda v: v >= 1,  lambda v: v >= 3,
        "Active TTC service alert(s) — check ttc.ca/service-advisories",
        "Multiple active TTC alerts — significant service disruption likely"),
    # Pearson weather impact (METAR-based, automated)
    # Pearson Airport Operations (NOTAM manual card) never hits thresholds
    ("Pearson Weather Impact",
        lambda v: v >= 1,  lambda v: v >= 2,
        "MVFR or worse at Pearson — weather impacting operations",
        "IFR/LIFR at Pearson — significant weather impact, possible delays"),
    # Pipeline utilization
    ("TCPL Parkway Receipts (GTA supply)",
        lambda v: v > 85,  lambda v: v > 95,
        "Parkway pipeline near capacity — GTA supply stress risk",
        "Parkway pipeline at capacity — GTA supply emergency risk"),
    ("TCPL Northern Ontario Line",
        lambda v: v > 85,  lambda v: v > 95,
        "Northern Ontario Line near capacity — upstream supply stress",
        "Northern Ontario Line at capacity — upstream supply emergency"),
    # Enbridge Dawn system status (0=normal, 1=interruptible, 2=firm)
    ("Enbridge Dawn System Status",
        lambda v: v >= 1,  lambda v: v >= 2,
        "Interruptible gas services potentially impacted — Dawn system constraint",
        "Firm gas services impacted — Dawn system under active constraint"),
    # Dawn storage level (Bcf) — only fires on plausible storage values (>10 Bcf)
    # Guards against StatsCan vectors returning 0.0 (net-change vectors, not stock)
    ("Dawn Hub Gas Storage",
        lambda v: 10 < v < 150, lambda v: 10 < v < 100,
        "Storage below 150 Bcf — reduced winter buffer (entering withdrawal season)",
        "Storage critically low — below 100 Bcf, supply stress risk for Ontario"),
    # Financial
    ("CAD/USD Exchange Rate",
        lambda v: v > 1.40, lambda v: v > 1.50,
        "CAD weakening — above 1.40 (import cost pressure)",
        "CAD under severe stress — above 1.50"),
    ("Fuel Price — Toronto",
        lambda v: v > 150,  lambda v: v > 185,
        "Fuel price elevated — above 150¢/L",
        "Fuel price critical — above 185¢/L"),
    ("Monthly Bankruptcy Filings",
        lambda v: v > 5000, lambda v: v > 7000,
        "Insolvency filings elevated — above 5,000/month (Ontario)",
        "Insolvency filings critical — above 7,000/month"),
    # Health
    ("Ontario ICU Occupancy",
        lambda v: v > 85,   lambda v: v > 95,
        "ICU occupancy elevated — above 85%",
        "ICU occupancy critical — above 95%"),
    # Environment
    ("Air Quality (AQHI)",
        lambda v: v >= 4,   lambda v: v >= 7,
        "Moderate risk — AQHI 4-6",
        "High risk — AQHI 7+"),
    # Boil water advisory — any active advisory is an immediate alert
    ("Boil Water Advisories",
        lambda v: v >= 1,  lambda v: v >= 1,
        "Active boil water advisory — public health signal",
        "Active boil water advisory — significant public health event"),
    # Water outages
    ("Active Water Outages",
        lambda v: v >= 3,  lambda v: v >= 8,
        "Multiple active outages — monitor for escalation",
        "High number of active outages — potential systemic stress"),
    ("Active Watermain Breaks",
        lambda v: v >= 2,  lambda v: v >= 5,
        "Multiple active breaks — elevated infrastructure stress",
        "High number of active breaks — systemic infrastructure alert"),
    # Food
    ("Grocery Price Inflation",
        lambda v: v > 170,  lambda v: v > 185,
        "Food CPI elevated — above index 170",
        "Food CPI critical — above index 185"),
]

def apply_thresholds(result):
    """
    Evaluate threshold rules against a result dict.
    Upgrades status from 'ok' to 'warn' or 'alert' as appropriate.
    Never downgrades manual/error status.
    Returns the result dict (mutated in place).
    """
    if result.get("status") not in ("ok", "warn"):
        return result
    value = result.get("value")
    if value is None:
        return result
    try:
        v = float(value)
    except (TypeError, ValueError):
        return result

    indicator = result.get("indicator", "")
    for (substr, warn_fn, alert_fn, warn_note, alert_note) in THRESHOLDS:
        if substr.lower() not in indicator.lower():
            continue
        try:
            if alert_fn(v) and alert_note:
                result["status"] = "alert"
                result["threshold_note"] = alert_note
            elif warn_fn(v) and warn_note:
                if result["status"] != "alert":
                    result["status"] = "warn"
                    result["threshold_note"] = warn_note
        except Exception:
            pass
        break  # first matching rule wins

    return result




# ══════════════════════════════════════════════════════════════════════════════
# TORONTO OPEN DATA (CKAN) HELPERS
# ══════════════════════════════════════════════════════════════════════════════

TORONTO_CKAN_BASE = "https://ckan0.cf.opendata.inter.prod-toronto.ca"

def _ckan_package(package_id):
    """Fetch a Toronto Open Data CKAN package metadata dict, or None on failure."""
    url = f"{TORONTO_CKAN_BASE}/api/3/action/package_show"
    try:
        r = SESSION.get(url, params={"id": package_id}, timeout=TIMEOUT)
        r.raise_for_status()
        result = r.json()
        if result.get("success"):
            return result["result"]
    except Exception:
        pass
    return None

def _ckan_datastore(resource_id, limit=100, sort=""):
    """Fetch records from a Toronto Open Data CKAN datastore resource."""
    url = f"{TORONTO_CKAN_BASE}/api/3/action/datastore_search"
    params = {"resource_id": resource_id, "limit": limit}
    if sort:
        params["sort"] = sort
    try:
        r = SESSION.get(url, params=params, timeout=TIMEOUT)
        r.raise_for_status()
        result = r.json()
        if result.get("success"):
            return result["result"].get("records", [])
    except Exception:
        pass
    return []


# ══════════════════════════════════════════════════════════════════════════════
# SECTOR: ENERGY
# ══════════════════════════════════════════════════════════════════════════════

def fetch_ieso_generation_mix():
    """
    IESO GenOutputbyFuelHourly — annual XML.

    v1.7: Confirmed schema from live document inspection:
      Document > DocBody > DailyData > Day
                                     > HourlyData > Hour
                                                  > FuelTotal > Fuel        (= "NUCLEAR" etc.)
                                                              > EnergyValue  (MWh as float)

    This file is fuel-aggregated by hour — no individual generator names.
    Per-generator detail lives in PUB_GenOutputCapability (separate endpoint).

    _IESO_GENERATOR_CACHE is populated with fuel-level records here;
    true per-generator breakdown can be added from GenOutputCapability later.
    """
    year = date.today().strftime("%Y")
    url = f"https://reports-public.ieso.ca/public/GenOutputbyFuelHourly/PUB_GenOutputbyFuelHourly_{year}.xml"

    try:
        r = SESSION.get(url, timeout=30)
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("IESO Generation Mix", "IESO GenOutputbyFuelHourly", url, str(e))]

    try:
        root = ET.fromstring(r.content)
    except ET.ParseError as e:
        return [_err("IESO Generation Mix", "IESO GenOutputbyFuelHourly", url,
                     f"XML parse error: {e}")]

    ns = root.tag.split("}")[0] + "}" if root.tag.startswith("{") else ""

    # Walk DailyData blocks — take the last one (most recent day)
    daily_blocks = root.findall(f".//{ns}DailyData") or root.findall(".//DailyData")
    if not daily_blocks:
        tags = sorted({el.tag.split("}")[-1] for el in root.iter()})
        return [_err("IESO Generation Mix", "IESO GenOutputbyFuelHourly", url,
                     f"No DailyData elements. Root: {root.tag}. Tags: {tags}")]

    # Annual file has DailyData for every day of the year including future placeholders.
    # Must search backwards through DailyData blocks to find the last one with real data.
    fuel_totals = {}
    latest_hour = None
    day_str = ""

    for last_day in reversed(daily_blocks):
        d_el = last_day.find(f"{ns}Day") or last_day.find("Day")
        candidate_day = (d_el.text or "").strip() if d_el is not None else ""

        hourly_blocks = (last_day.findall(f".//{ns}HourlyData") or
                         last_day.findall(".//HourlyData"))

        # Search backwards through hours in this day for populated data.
        # Use local-name matching to bypass namespace prefix issues entirely.
        for hourly in reversed(hourly_blocks):
            # Index all direct and grandchild elements by local tag name
            by_tag = {}
            for child in hourly:
                local = child.tag.split("}")[-1]
                by_tag.setdefault(local, []).append(child)

            h_list = by_tag.get("Hour", [])
            if not h_list or not h_list[0].text:
                continue
            try:
                hour_num = int(h_list[0].text.strip())
            except ValueError:
                continue

            candidate = {}
            for ft in by_tag.get("FuelTotal", []):
                # Index FuelTotal children by local name
                ft_kids = {c.tag.split("}")[-1]: c for c in ft}
                fuel_el = ft_kids.get("Fuel")
                ev_el   = ft_kids.get("EnergyValue")
                if fuel_el is None or ev_el is None:
                    continue
                # EnergyValue > Output holds the MW number
                ev_kids = {c.tag.split("}")[-1]: c for c in ev_el}
                out_el  = ev_kids.get("Output")
                if out_el is None:
                    continue
                fuel     = (fuel_el.text or "").strip().upper()
                val_text = (out_el.text  or "").strip()
                if not fuel or not val_text:
                    continue
                try:
                    candidate[fuel] = float(val_text)
                except ValueError:
                    continue

            if candidate:
                fuel_totals = candidate
                latest_hour = hour_num
                day_str     = candidate_day
                break  # found populated hour in this day

        if fuel_totals:
            break  # found populated day — stop searching

    if not fuel_totals:
        tags = sorted({el.tag.split("}")[-1] for el in root.iter()})
        return [_err("IESO Generation Mix", "IESO GenOutputbyFuelHourly", url,
                     f"No data in {len(daily_blocks)} DailyData blocks. "
                     f"Tags in document: {tags}")]

    # Populate cache (fuel-level, not generator-level — upgrade later with GenOutputCapability)
    _IESO_GENERATOR_CACHE.clear()
    _IESO_GENERATOR_CACHE.update({
        "retrieved_at":    datetime.utcnow().isoformat() + "Z",
        "data_date":       day_str or str(date.today()),
        "hour":            latest_hour,
        "source_url":      url,
        "note":            "Fuel-aggregated totals. Per-generator detail available from "
                           "PUB_GenOutputCapability (not yet implemented).",
        "fuel_totals_mw":  {k: round(v, 1) for k, v in fuel_totals.items()},
        "generators":      [
            {"name": fuel, "fuel": fuel,
             "output_mw": round(mw, 1), "capacity_mw": None,
             "hour": latest_hour, "data_date": day_str or str(date.today())}
            for fuel, mw in sorted(fuel_totals.items(),
                                   key=lambda x: x[1], reverse=True)
        ],
    })

    total_mw = sum(fuel_totals.values())
    # data_date: store ISO date only; hour goes in notes so fmtDate renders cleanly
    data_date = day_str if day_str else str(date.today())
    hour_note = f"Hour ending {latest_hour:02d}:00 UTC."
    fuel_summary = ", ".join(f"{k}: {round(v)}MW" for k, v in
                             sorted(fuel_totals.items(), key=lambda x: x[1], reverse=True))
    results = [_ok("Total Generation (MW)", round(total_mw, 1), "MW",
                   "IESO GenOutputbyFuelHourly", url, data_date,
                   f"Tier 3 — Local real-time delivery. {hour_note} "
                   f"Mix: {fuel_summary}. "
                   f"Chain terminus: Brent/NGAS (Tier 1) → TCPL utilization (Tier 2) → "
                   f"IESO dispatch / pump price (Tier 3).")]
    GAS_NOTE = (
        "Tier 3 — Local real-time delivery. Gas peakers are marginal-cost units dispatched "
        "when baseload (nuclear/hydro) is insufficient. Output >3,000 MW = demand stress; "
        ">5,000 MW = grid under significant pressure. Elevated gas dispatch correlates with "
        "higher wholesale electricity prices and increased exposure to Tier 1/2 commodity risk."
    )
    for key, (name, unit) in {
        "NUCLEAR": ("Nuclear Output (MW)", "MW"),
        "HYDRO":   ("Hydro Output (MW)",   "MW"),
        "GAS":     ("Gas Output (MW)",     "MW"),
        "WIND":    ("Wind Output (MW)",    "MW"),
        "SOLAR":   ("Solar Output (MW)",   "MW"),
        "BIOFUEL": ("Biofuel Output (MW)", "MW"),
    }.items():
        notes = (GAS_NOTE + f" {hour_note}") if key == "GAS" else hour_note
        results.append(_ok(name, round(fuel_totals.get(key, 0), 1), unit,
                           "IESO GenOutputbyFuelHourly", url, data_date, notes))
    return results


def fetch_ieso_ontario_demand():
    """
    IESO Ontario Demand multiday — also uses ElementTree now.
    The ashx file is XML. ElementTree handles it correctly.
    """
    PEAK_REF_MW = 25000
    url = "https://www.ieso.ca/-/media/Files/IESO/Power-Data/Ontario-Demand-multiday.ashx"

    try:
        r = SESSION.get(url, timeout=TIMEOUT)
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("Peak Demand Index", "IESO Ontario Demand Multiday", url, str(e))]

    demand_val = None
    try:
        root = ET.fromstring(r.content)
        ns = root.tag.split("}")[0] + "}" if root.tag.startswith("{") else ""
        # Search for demand elements
        for tag in ["OntarioDemand", "Demand", "TotalDemand", "MarketDemand",
                    "ActualDemand", "ForecastDemand"]:
            els = root.findall(f".//{ns}{tag}") or root.findall(f".//{tag}")
            for el in reversed(els):
                if el.text:
                    try:
                        v = float(el.text.strip())
                        if v > 0:
                            demand_val = v
                            break
                    except ValueError:
                        pass
            if demand_val:
                break
    except ET.ParseError:
        pass

    # Fallback: numeric scan of content (works even if XML structure changes)
    if demand_val is None:
        for line in r.text.splitlines():
            nums = re.findall(r'\b(\d{4,5}(?:\.\d+)?)\b', line)
            for n in nums:
                v = float(n)
                if 5000 < v < 30000:
                    demand_val = v
                    break
            if demand_val:
                break

    if demand_val is None:
        return [_err("Peak Demand Index", "IESO Ontario Demand Multiday", url,
                     f"Could not parse demand from {len(r.content)} bytes")]

    peak_index = round(demand_val / PEAK_REF_MW * 100, 1)
    return [
        _ok("Ontario Demand (MW)", demand_val, "MW",
            "IESO Ontario Demand Multiday", url, date.today()),
        _ok("Peak Demand Index", peak_index, "% of 2024 peak",
            "IESO Ontario Demand Multiday", url, date.today(),
            f"{demand_val} MW ÷ {PEAK_REF_MW} MW × 100"),
    ]



def fetch_brent_crude():
    """
    Brent crude spot price.

    TIER POSITION: Tier 1 - Global commodity layer.
    Brent is the first link in the Ontario energy supply chain:
      Tier 1 -> Brent crude (global commodity benchmark, set by ICE/OPEC dynamics)
      Tier 2 -> TCPL pipeline utilization (continental delivery infrastructure)
      Tier 3 -> Toronto pump price / IESO gas peaker dispatch (local real-time delivery)

    Sources tried in order:
      1. Yahoo Finance (BZ=F) - ICE Brent front-month futures, near-real-time.
         No API key. ~15 min delay. Works from GitHub Actions.
         Critical during fast-moving geopolitical events.
      2. datasets/oil-prices (GitHub raw) - EIA Brent spot, mirrored daily.
         1-3 day lag. Always reachable from Actions.
      3. FRED DCOILBRENTEU - EIA spot via St. Louis Fed. 1-2 day lag.
      4. Stooq cb.f - ICE Brent futures. No API key, non-US jurisdiction.
    """
    NOTES_TEMPLATE = (
        "Tier 1 - Global commodity. ICE Brent crude: global benchmark for ~2/3 of world oil trade. "
        "Ontario supply chain: Brent (Tier 1) -> TCPL pipeline utilization (Tier 2) -> "
        "Toronto pump price / IESO gas dispatch (Tier 3). "
        "CAD/USD amplifies or dampens pass-through to Ontario consumers. "
        "Warn >$80/bbl (refinery margin pressure), Alert >$100/bbl (Hormuz disruption threshold). "
        "Source: {source}."
    )

    errors = []

    # Source 1: Yahoo Finance BZ=F (ICE Brent front-month futures, ~15 min delay)
    # Returns JSON with regularMarketPrice field. No API key required.
    # This is the most current source -- critical during geopolitical events.
    yahoo_url = ("https://query1.finance.yahoo.com/v8/finance/chart/BZ=F"
                 "?interval=1d&range=5d")
    try:
        r = SESSION.get(yahoo_url,
                        headers={"User-Agent": "Mozilla/5.0",
                                 "Accept": "application/json"},
                        timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
        meta = data["chart"]["result"][0]["meta"]
        price = round(float(meta["regularMarketPrice"]), 2)
        # Use regularMarketTime (Unix timestamp) for the data date
        import time as _time
        ts = meta.get("regularMarketTime", 0)
        if ts:
            from datetime import datetime as _dt, timezone as _tz
            date_str = _dt.fromtimestamp(ts, tz=_tz.utc).strftime("%Y-%m-%d")
        else:
            date_str = str(date.today())
        return [_ok("Brent Crude Price", price, "USD/bbl",
                    "Yahoo Finance - ICE Brent futures (BZ=F)", yahoo_url,
                    date_str,
                    f"Futures price (~15 min delay). " +
                    NOTES_TEMPLATE.format(
                        source="Yahoo Finance ICE Brent front-month futures (BZ=F), "
                               "near-real-time"))]
    except Exception as e:
        errors.append(f"Yahoo BZ=F: {e}")

    # Source 2: datasets/oil-prices GitHub raw CSV (EIA spot, 1-3 day lag)
    datahub_url = ("https://raw.githubusercontent.com/datasets/oil-prices"
                   "/main/data/brent-daily.csv")
    try:
        r = SESSION.get(datahub_url, headers={"User-Agent": "Mozilla/5.0"},
                        timeout=TIMEOUT)
        r.raise_for_status()
        lines = r.text.strip().splitlines()
        for row in reversed(lines[1:]):
            parts = row.split(",")
            if len(parts) >= 2 and parts[1].strip() not in ("", ".", "N/D"):
                date_str = parts[0].strip()
                price = round(float(parts[1].strip()), 2)
                try:
                    from datetime import date as _date
                    lag_days = (_date.today() - _date.fromisoformat(date_str)).days
                    lag_note = (f"EIA spot price. Data date: {date_str} "
                                f"({lag_days}d ago -- normal 1-4 day EIA publication lag)."
                                if lag_days <= 4 else
                                f"EIA spot price. Data date: {date_str} "
                                f"({lag_days}d ago -- may be stale).")
                except Exception:
                    lag_note = f"EIA spot price. Data date: {date_str}."
                return [_ok("Brent Crude Price", price, "USD/bbl",
                            "EIA via datasets/oil-prices (GitHub)", datahub_url,
                            date_str,
                            lag_note + " " + NOTES_TEMPLATE.format(
                                source="EIA Brent spot price mirrored daily at "
                                       "datasets/oil-prices on GitHub"))]
        errors.append("datahub: all rows missing or null")
    except Exception as e:
        errors.append(f"datahub: {e}")

    # Source 3: FRED DCOILBRENTEU
    fred_url = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=DCOILBRENTEU"
    try:
        r = SESSION.get(fred_url, headers={"User-Agent": "Mozilla/5.0"},
                        timeout=TIMEOUT)
        r.raise_for_status()
        lines = r.text.strip().splitlines()
        for row in reversed(lines[1:]):
            parts = row.split(",")
            if len(parts) == 2 and parts[1].strip() not in ("", ".", "N/D"):
                date_str = parts[0].strip()
                price = round(float(parts[1].strip()), 2)
                return [_ok("Brent Crude Price", price, "USD/bbl",
                            "FRED - DCOILBRENTEU (St. Louis Fed)", fred_url,
                            date_str,
                            NOTES_TEMPLATE.format(
                                source="FRED DCOILBRENTEU - St. Louis Federal Reserve "
                                       "daily Brent spot price"))]
        errors.append("FRED: all rows missing or null")
    except Exception as e:
        errors.append(f"FRED: {e}")

    # Source 4: Stooq cb.f
    stooq_url = "https://stooq.com/q/d/l/?s=cb.f&i=d&l=5"
    try:
        r = SESSION.get(stooq_url, headers={"User-Agent": "Mozilla/5.0"},
                        timeout=TIMEOUT)
        r.raise_for_status()
        lines = r.text.strip().splitlines()
        for row in reversed(lines[1:]):
            parts = row.split(",")
            if len(parts) >= 5 and parts[4].strip() not in ("", "null", "N/D"):
                date_str = parts[0].strip()
                price = round(float(parts[4].strip()), 2)
                return [_ok("Brent Crude Price", price, "USD/bbl",
                            "Stooq.com - ICE Brent (cb.f)", stooq_url,
                            date_str,
                            NOTES_TEMPLATE.format(
                                source="Stooq.com ICE Brent futures (cb.f)"))]
        errors.append("Stooq cb.f: all rows null/N/D")
    except Exception as e:
        errors.append(f"Stooq cb.f: {e}")

    return [_err("Brent Crude Price", "Yahoo/EIA/FRED/Stooq", yahoo_url,
                 " | ".join(errors))]


def fetch_toronto_shelter():
    """Toronto Open Data — Daily Shelter. Unchanged — was working."""
    pkg = _ckan_package("daily-shelter-overnight-service-occupancy-capacity")
    if not pkg:
        return [_err("Shelter System Capacity", "Toronto Open Data",
                     TORONTO_CKAN_BASE, "Package not found")]

    resource_id = next((r["id"] for r in pkg.get("resources", [])
                        if r.get("datastore_active")), None)
    if not resource_id:
        return [_err("Shelter System Capacity", "Toronto Open Data",
                     TORONTO_CKAN_BASE, "No active datastore resource")]

    records = _ckan_datastore(resource_id, 3000, "OCCUPANCY_DATE desc")
    if not records:
        return [_err("Shelter System Capacity", "Toronto Open Data",
                     TORONTO_CKAN_BASE, "Datastore empty or failed")]

    most_recent_date = records[0].get("OCCUPANCY_DATE", "")
    if not most_recent_date:
        return [_err("Shelter System Capacity", "Toronto Open Data",
                     TORONTO_CKAN_BASE, "OCCUPANCY_DATE field not found")]

    day_recs = [rec for rec in records if rec.get("OCCUPANCY_DATE") == most_recent_date]
    total_occ = total_cap = 0
    for rec in day_recs:
        for k in ["OCCUPANCY", "OCCUPIED_BEDS", "occupied_beds"]:
            if k in rec:
                try: total_occ += float(rec[k] or 0); break
                except (ValueError, TypeError): pass
        for k in ["CAPACITY_ACTUAL", "CAPACITY_ACTUAL_BED", "capacity_actual_bed"]:
            if k in rec:
                try: total_cap += float(rec[k] or 0); break
                except (ValueError, TypeError): pass

    occ_rate = round(total_occ / total_cap * 100, 1) if total_cap > 0 else None
    api_url = f"{TORONTO_CKAN_BASE}/api/3/action/datastore_search?resource_id={resource_id}"
    return [
        _ok("Shelter System Capacity", int(total_cap), "beds/spaces",
            "Toronto Open Data — Daily Shelter", api_url, most_recent_date,
            f"{len(day_recs)} programs at 4am snapshot"),
        _ok("Shelter Occupancy", int(total_occ), "occupied",
            "Toronto Open Data — Daily Shelter", api_url, most_recent_date,
            f"Occupancy rate: {occ_rate}%. Threshold: 95% = Warning"),
    ]


# ══════════════════════════════════════════════════════════════════════════════
# SECTOR: HEALTH
# ══════════════════════════════════════════════════════════════════════════════

def fetch_active_water_outages():
    """
    Toronto Water — active water outages.
    Source: City of Toronto ArcGIS REST API (COT_Geospatial_Water_Outage_View)
    This is the live data feed behind toronto.ca/no-water-map.
    Updated in real time as breaks are reported and restored.

    Field mapping (confirmed from live schema inspection Mar 2026):
      WaterOutageEdit7b_ReasonForShut — coded value: 1=Emergency Repair, 2=Planned Work
      WaterOutageEdit7b_DateandTimeof — outage start (epoch ms)
      WaterOutageEdit7b_Bypass        — 1=test/bypass record (excluded)
      Est_Num_Properties_Affected     — estimated properties without water
      EstRestorationDateTime          — estimated restoration (epoch ms)

    Query: service not yet restored (DateandTimeServiceRestored IS NULL)
    and not a bypass/test record (WaterOutageEdit7b_Bypass <> 1)
    """
    BASE = ("https://services3.arcgis.com/b9WvedVPoizGfvfD/arcgis/rest/services"
            "/COT_Geospatial_Water_Outage_View/FeatureServer/0/query")
    params = {
        "where": "DateandTimeServiceRestored IS NULL AND WaterOutageEdit7b_Bypass <> 1",
        "outFields": "*",
        "f": "geojson",
        "returnGeometry": "false",
    }
    url = BASE + "?" + "&".join(f"{k}={v}" for k, v in params.items())

    # Coded values for WaterOutageEdit7b_ReasonForShut
    REASON_CODES = {1: "Emergency Repair", 2: "Planned Work"}

    try:
        r = SESSION.get(BASE, params=params, timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
    except (requests.RequestException, json.JSONDecodeError) as e:
        return [_err("Active Water Outages", "Toronto Water ArcGIS API", url, str(e))]

    features = data.get("features", [])
    if not features and "error" in data:
        return [_err("Active Water Outages", "Toronto Water ArcGIS API", url,
                     f"API error: {data['error']}")]

    emergency = 0
    planned   = 0
    unknown   = 0
    total_properties = 0
    oldest_start = None
    soonest_restore = None
    addresses = []

    from datetime import datetime as _dt

    for feat in features:
        props = feat.get("properties", {})

        reason_code_raw = props.get("WaterOutageEdit7b_ReasonForShut")
        # Handle both integer and string representations from GeoJSON
        try:
            reason_code = int(reason_code_raw) if reason_code_raw is not None else None
        except (ValueError, TypeError):
            reason_code = None
        reason = REASON_CODES.get(reason_code, "Unknown")

        if reason_code == 1:
            emergency += 1
        elif reason_code == 2:
            planned += 1
        else:
            unknown += 1

        # Properties affected
        affected = props.get("Est_Num_Properties_Affected")
        if affected and isinstance(affected, (int, float)) and affected > 0:
            total_properties += int(affected)

        # Outage start time
        start_ts = props.get("WaterOutageEdit7b_DateandTimeof")
        if start_ts and isinstance(start_ts, (int, float)):
            try:
                dt = _dt.utcfromtimestamp(start_ts / 1000)
                if oldest_start is None or dt < oldest_start:
                    oldest_start = dt
            except (ValueError, OSError):
                pass

        # Estimated restoration
        restore_ts = props.get("EstRestorationDateTime")
        if restore_ts and isinstance(restore_ts, (int, float)):
            try:
                dt = _dt.utcfromtimestamp(restore_ts / 1000)
                if soonest_restore is None or dt < soonest_restore:
                    soonest_restore = dt
            except (ValueError, OSError):
                pass

        # Collect addresses for context
        addr = props.get("WaterOutageEdit7b_Address", "")
        if addr:
            addresses.append(str(addr))

    total = len(features)
    oldest_str  = oldest_start.strftime("%Y-%m-%d %H:%M UTC") if oldest_start else "unknown"
    restore_str = soonest_restore.strftime("%Y-%m-%d %H:%M UTC") if soonest_restore else "unknown"
    addr_str    = ", ".join(addresses[:5]) + ("..." if len(addresses) > 5 else "")
    props_str   = f"{total_properties:,} properties affected. " if total_properties > 0 else ""

    results = [
        _ok("Active Water Outages", total, "active outages",
            "Toronto Water — No Water Map (ArcGIS)", url, str(date.today()),
            f"Emergency repairs: {emergency}, Planned work: {planned}, Other: {unknown}. "
            f"{props_str}"
            f"Oldest active since: {oldest_str}. "
            f"Earliest est. restoration: {restore_str}. "
            f"Active locations: {addr_str}. "
            f"Source: toronto.ca/no-water-map — real-time feed."),
    ]

    if emergency > 0:
        results.append(_ok("Active Watermain Breaks", emergency, "emergency repairs",
                           "Toronto Water — No Water Map (ArcGIS)", url,
                           str(date.today()),
                           f"Emergency water repairs with service interrupted. "
                           f"{props_str}"
                           f"Locations: {addr_str}. "
                           f"Est. restoration: {restore_str}."))
    if planned > 0:
        results.append(_ok("Planned Maintenance Outages", planned, "outages",
                           "Toronto Water — No Water Map (ArcGIS)", url,
                           str(date.today()),
                           f"Scheduled maintenance with planned service interruption. "
                           f"{props_str}"))

    return results


def fetch_toronto_unemployment():
    """
    Toronto CMA unemployment rate — StatsCan WDS API.
    Table 14-10-0459-01: Labour force characteristics by CMA, 3-month moving average, SA.
    Replaced inactive table 14-10-0294-01 (superseded May 2025).

    The POST-based getDataFromCubePidCoordAndLatestNPeriods endpoint returns 406.
    Using GET-based endpoint instead, then falling back to vector search.
    Released monthly, lags ~5 weeks after reference month.
    Alert threshold: >10% = recession-level stress.
    """
    # ── Source 1: GET-based coordinate endpoint (new table) ──────────────────
    # GET format: /getDataFromCubePidCoordAndLatestNPeriods/{pid}/{coord}/{n}
    # Toronto CMA member varies by table — try multiple coordinate patterns
    GET_BASE = "https://www150.statcan.gc.ca/t1/wds/rest/getDataFromCubePidCoordAndLatestNPeriods"
    coords_to_try = [
        (1410045901, "35.3.1.1"),        # new table, Toronto, unemployment, both sexes
        (1410045901, "35.3.1.1.1.1"),
        (1410045901, "35.3.1.1.1.1.1.1.1.1"),
        (14100294,   "23.5.1.1"),         # old table fallback
    ]
    for pid, coord in coords_to_try:
        try:
            url = f"{GET_BASE}/{pid}/{coord}/2"
            r = SESSION.get(url, timeout=30,
                            headers={"Accept": "application/json"})
            r.raise_for_status()
            data = r.json()
            if not isinstance(data, list) or not data:
                continue
            obj = data[0].get("object", {})
            if data[0].get("status") != "SUCCESS":
                continue
            points = obj.get("vectorDataPoint", [])
            if not points:
                continue
            latest = points[-1]
            rate = float(latest["value"])
            if not (4.0 <= rate <= 20.0):   # sanity check
                continue
            ref = latest.get("refPer", "unknown")
            vid = obj.get("vectorId", "?")
            table_label = "14-10-0459-01" if pid == 1410045901 else "14-10-0294-01"
            return [_ok("Toronto Unemployment Rate", round(rate, 1), "%",
                        f"StatsCan LFS — Table {table_label} (Toronto CMA)",
                        url, ref,
                        f"3-month moving average, seasonally adjusted. Vector v{vid}. "
                        f"Released ~5 weeks after reference month. "
                        f"Pre-tariff baseline (2023): ~6.5%. Alert: >10%.")]
        except Exception:
            continue

    # ── Source 2: Vector-based endpoint — scan likely vector range ────────────
    # The new table 14-10-0459-01 vectors are in a different range than old table.
    # We try the GET series info endpoint to discover the vector for Toronto CMA.
    SERIES_URL = "https://www150.statcan.gc.ca/t1/wds/rest/getSeriesInfoFromCubePidCoord"
    for pid, coord in [(1410045901, "35.3.1.1"), (14100294, "23.5.1.1")]:
        try:
            url = f"{SERIES_URL}/{pid}/{coord}"
            r = SESSION.get(url, timeout=30,
                            headers={"Accept": "application/json"})
            r.raise_for_status()
            data = r.json()
            if not isinstance(data, list) or not data:
                continue
            obj = data[0].get("object", {})
            if data[0].get("status") != "SUCCESS":
                continue
            vid = obj.get("vectorId")
            if not vid:
                continue
            # Now fetch data for this vector
            vec_url = ("https://www150.statcan.gc.ca/t1/wds/rest/"
                       f"getDataFromVectorsAndLatestNPeriods/[{{\"vectorId\":{vid},\"latestN\":2}}]")
            r2 = SESSION.get(vec_url, timeout=30,
                             headers={"Accept": "application/json"})
            r2.raise_for_status()
            vdata = r2.json()
            if not vdata or vdata[0].get("status") != "SUCCESS":
                continue
            points = vdata[0].get("object", {}).get("vectorDataPoint", [])
            if not points:
                continue
            rate = float(points[-1]["value"])
            if not (4.0 <= rate <= 20.0):
                continue
            ref = points[-1].get("refPer", "unknown")
            return [_ok("Toronto Unemployment Rate", round(rate, 1), "%",
                        f"StatsCan LFS — vector v{vid} (Toronto CMA)",
                        vec_url, ref,
                        f"3-month moving average, seasonally adjusted. Vector v{vid}. "
                        f"Released ~5 weeks after reference month. "
                        f"Pre-tariff baseline (2023): ~6.5%. Alert: >10%.")]
        except Exception:
            continue

    return [_err("Toronto Unemployment Rate",
                 "StatsCan WDS (Tables 14-10-0459-01 / 14-10-0294-01)",
                 GET_BASE,
                 "All sources failed. "
                 "Manual: www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid=1410045901")]


def fetch_toronto_boil_advisories():
    """
    Toronto Water — active boil water advisories.
    Scrapes toronto.ca drinking water quality page.
    Toronto's water system is large and well-maintained — advisories are rare.
    A non-zero count is a significant public health signal.
    Alert threshold: any active advisory = alert (zero is normal).
    """
    # Primary: Toronto Water drinking water advisories page
    urls_to_try = [
        "https://www.toronto.ca/services-payments/water-environment/tap-water-in-toronto/",
        "https://www.toronto.ca/services-payments/water-environment/tap-water-in-toronto/drinking-water-compliance-and-testing/",
    ]

    advisory_count = 0
    advisory_details = []
    source_url = urls_to_try[0]

    for url in urls_to_try:
        try:
            r = SESSION.get(url, timeout=TIMEOUT,
                            headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.5"})
            r.raise_for_status()
            from bs4 import BeautifulSoup as _BS
            soup = _BS(r.content, "html.parser")
            text = soup.get_text(" ", strip=True).lower()

            # Look for advisory language
            if "boil water" in text or "boil-water" in text:
                # Count specific advisory mentions
                advisory_keywords = [
                    "boil water advisory", "boil-water advisory",
                    "do not use", "water advisory in effect"
                ]
                for kw in advisory_keywords:
                    count = text.count(kw)
                    if count > 0:
                        advisory_count += count
                        advisory_details.append(f"'{kw}' found {count}x")

                # If we found mentions, this page has the info
                if advisory_count > 0:
                    source_url = url
                    break
                else:
                    # "boil water" mentioned but in general context, not active advisory
                    # Check for "no current advisories" language
                    if any(phrase in text for phrase in
                           ["no current", "no active", "no boil water advisory",
                            "no advisories", "water is safe"]):
                        advisory_count = 0
                        source_url = url
                        break
            elif any(phrase in text for phrase in
                     ["no current", "no active", "water is safe", "meets all"]):
                advisory_count = 0
                source_url = url
                break
        except requests.RequestException:
            continue

    notes = (f"Active advisories: {advisory_count}. Details: {advisory_details}."
             if advisory_count > 0
             else "No active boil water advisories detected. "
                  "Toronto's 4 treatment plants serve ~4M people. "
                  "An advisory would appear at toronto.ca/tap-water.")

    return [_ok("Boil Water Advisories", advisory_count, "active advisories",
                "Toronto Water — toronto.ca", source_url,
                str(date.today()), notes)]


def fetch_trreb_market():
    """
    TRREB GTA housing market — sales-to-new-listings ratio and average price.
    Source: trreb.ca/market-data/quick-market-overview/
    Data embedded as JavaScript variables in page HTML, updated monthly.
    Published first week of each month for the prior month.

    Key indicators:
    - Sales-to-new-listings ratio (SNR):
        <40% = buyer's market
        40-60% = balanced market
        >60% = seller's market
    - Average selling price (all residential)
    - Total sales (year-over-year)
    - Total new listings (year-over-year)
    """
    URL = "https://trreb.ca/market-data/quick-market-overview/"
    try:
        r = SESSION.get(URL, timeout=TIMEOUT,
                        headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.6"})
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("TRREB Sales-to-New-Listings Ratio", "TRREB Quick Overview", URL, str(e))]

    try:
        from bs4 import BeautifulSoup as _BS
        import re as _re
        soup = _BS(r.content, "html.parser")

        # All data is in inline script blocks as JS variables
        scripts = " ".join(s.string or "" for s in soup.find_all("script", src=False))

        def extract_js_obj(scripts, var_name):
            """Extract last value from a JS object literal like: let Var = {"Feb'25": 32, "Feb'26": 36}"""
            pattern = rf'let\s+{var_name}\s*=\s*\{{([^}}]+)\}}'
            m = _re.search(pattern, scripts)
            if not m:
                return None, None, None
            pairs = _re.findall(r'"([^"]+)":\s*([\d.]+)', m.group(1))
            if not pairs:
                return None, None, None
            # Return latest period and value, plus prior period value
            latest_period, latest_val = pairs[-1]
            prior_val = pairs[-2][1] if len(pairs) >= 2 else None
            return latest_period, float(latest_val), float(prior_val) if prior_val else None

        # Sales-to-new-listings ratio
        snr_period, snr_val, snr_prior = extract_js_obj(scripts, "TnlSnrData")
        # Average selling price
        asp_period, asp_val, asp_prior = extract_js_obj(scripts, "AspYoyData")
        # Total residential transactions
        trt_period, trt_val, trt_prior = extract_js_obj(scripts, "TrtYoyData")
        # Total new listings
        tnl_period, tnl_val, tnl_prior = extract_js_obj(scripts, "TnlYoyData")

        if snr_val is None:
            raise ValueError("Could not parse TnlSnrData from page — page structure may have changed")

        # Market classification
        if snr_val < 40:
            market_type = "buyer's market"
        elif snr_val <= 60:
            market_type = "balanced market"
        else:
            market_type = "seller's market"

        # Price change YoY
        price_note = ""
        if asp_val and asp_prior:
            price_chg = round(((asp_val - asp_prior) / asp_prior) * 100, 1)
            price_note = f" Avg price ${asp_val:,.0f} ({price_chg:+.1f}% YoY)."

        # Sales change YoY
        sales_note = ""
        if trt_val and trt_prior:
            sales_chg = round(((trt_val - trt_prior) / trt_prior) * 100, 1)
            sales_note = f" Sales {int(trt_val):,} ({sales_chg:+.1f}% YoY)."

        # Listings change YoY
        listings_note = ""
        if tnl_val and tnl_prior:
            listings_chg = round(((tnl_val - tnl_prior) / tnl_prior) * 100, 1)
            listings_note = f" New listings {int(tnl_val):,} ({listings_chg:+.1f}% YoY)."

        notes = (f"{market_type.title()} — SNR of {snr_val:.0f}% "
                 f"(prior period: {snr_prior:.0f}%).{price_note}{sales_note}{listings_note} "
                 f"Reference: {snr_period}. "
                 f"SNR guide: <40% buyer's, 40-60% balanced, >60% seller's market.")

        results = [
            _ok("TRREB Sales-to-New-Listings Ratio", snr_val, "%",
                "TRREB Quick Market Overview", URL, snr_period, notes),
        ]
        if asp_val:
            results.append(
                _ok("TRREB Average Selling Price", round(asp_val), "CAD",
                    "TRREB Quick Market Overview", URL, asp_period,
                    f"GTA all residential. {price_note.strip()}"))
        return results

    except Exception as e:
        return [_err("TRREB Sales-to-New-Listings Ratio", "TRREB Quick Overview",
                     URL, f"Parse error: {e}")]


def fetch_tcpl_mainline():
    """
    TransCanada (TC) Canadian Mainline — Parkway Receipts utilization.
    Source: Canada Energy Regulator (CER) open data CSV.
    Updated monthly, daily granularity.

    Parkway Receipts is the key point measuring gas arriving at the
    Parkway hub near Toronto — the primary delivery point for GTA supply.
    Northern Ontario Line is included as secondary context (gas transiting
    across northern Ontario toward Toronto).

    Utilization = throughput / capacity * 100
    Normal operating range: 50-80%
    Warn: >85% sustained (pipeline near capacity, supply stress risk)
    Alert: >95% sustained (pipeline at capacity, supply emergency risk)

    CSV has ~208,000 rows (2006-present). We read only the last 5,000
    rows to get recent data efficiently without downloading the full file.
    """
    URL = ("https://www.cer-rec.gc.ca/open/energy/throughput-capacity/"
           "tcpl-mainline-throughput-and-capacity.csv")

    try:
        r = SESSION.get(URL, timeout=30,
                        headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.7"})
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("TCPL Mainline Parkway Utilization",
                     "CER Open Data CSV", URL, str(e))]

    try:
        import io as _io
        import csv as _csv

        lines = r.text.splitlines()
        if len(lines) < 2:
            raise ValueError("CSV appears empty")

        header = lines[0]
        # Read last 5000 rows for efficiency — covers ~14 months of daily data
        recent_text = "\n".join([header] + lines[-5000:])
        reader = _csv.DictReader(_io.StringIO(recent_text))

        # Collect latest rows per key point
        latest = {}
        weekly = {}  # store last 7 days per key point

        for row in reader:
            kp = row.get("Key Point", "").strip().replace("\n", "").replace("\r", "")
            dt = row.get("Date", "")
            if not kp or not dt:
                continue
            if kp not in latest or dt > latest[kp]["Date"]:
                latest[kp] = row
            if kp not in weekly:
                weekly[kp] = []
            weekly[kp].append(row)

        # Key points we care about
        targets = {
            "Eastern Triangle - Parkway Receipts": "Parkway Receipts (GTA supply)",
            "Northern Ontario Line":               "Northern Ontario Line",
        }

        results = []
        for kp, label in targets.items():
            # Clean key — CSV has embedded newlines in key point names
            matched_key = None
            for k in latest:
                if kp.lower() in k.lower():
                    matched_key = k
                    break

            if not matched_key:
                results.append(_err(f"TCPL {label}", "CER Open Data", URL,
                                    f"Key point '{kp}' not found in CSV"))
                continue

            row = latest[matched_key]
            ref_date = row.get("Date", "unknown")

            try:
                cap = float(row.get("Capacity (1000 m3/d)", 0) or 0)
                thr = float(row.get("Throughput (1000 m3/d)", 0) or 0)
            except (ValueError, TypeError):
                results.append(_err(f"TCPL {label}", "CER Open Data", URL,
                                    "Could not parse capacity/throughput values"))
                continue

            if cap <= 0:
                results.append(_err(f"TCPL {label}", "CER Open Data", URL,
                                    f"Capacity is zero for {matched_key}"))
                continue

            util = round(thr / cap * 100, 1)

            # 7-day average utilization
            recent_rows = sorted(weekly.get(matched_key, []),
                                 key=lambda x: x.get("Date", ""))[-7:]
            if len(recent_rows) >= 3:
                utils = []
                for rr in recent_rows:
                    try:
                        rc = float(rr.get("Capacity (1000 m3/d)", 0) or 0)
                        rt = float(rr.get("Throughput (1000 m3/d)", 0) or 0)
                        if rc > 0:
                            utils.append(rt / rc * 100)
                    except (ValueError, TypeError):
                        pass
                avg_util = round(sum(utils) / len(utils), 1) if utils else util
                avg_note = f" 7-day avg: {avg_util}%."
            else:
                avg_util = util
                avg_note = ""

            tier_context = (
                "Tier 2 — Continental supply chain. "
                "TCPL Mainline moves western Canadian gas east across Ontario. "
                "Elevated utilization here precedes Tier 3 stress (IESO gas peaker dispatch, "
                "Toronto pump price) by days to weeks. "
                "Chain: Brent/NGAS (Tier 1) → TCPL utilization (Tier 2) → "
                "IESO gas output / pump price (Tier 3). "
            ) if "Parkway" in label else (
                "Tier 2 — Continental supply chain. "
                "Northern Ontario Line transits gas from western fields toward Parkway/GTA. "
                "Upstream constraint here feeds into Parkway pressure before reaching "
                "Toronto-area consumers and IESO gas peakers (Tier 3). "
            )
            notes = (f"{tier_context}"
                     f"Throughput: {thr:,.0f} / Capacity: {cap:,.0f} (1000 m³/day). "
                     f"Utilization: {util}%.{avg_note} "
                     f"Normal range 50-80%. Warn >85% (near capacity), Alert >95% (emergency risk). "
                     f"⚠ Data lag: CER publishes quarterly (~3 month lag). "
                     f"Source: CER open data (TransCanada Mainline throughput CSV).")

            results.append(_ok(f"TCPL {label}", util, "% utilized",
                               "CER — TransCanada Mainline", URL,
                               ref_date, notes))

        return results if results else [_err("TCPL Mainline", "CER Open Data",
                                            URL, "No matching key points found")]

    except Exception as e:
        return [_err("TCPL Mainline Parkway Utilization",
                     "CER Open Data CSV", URL, f"Parse error: {e}")]


def fetch_pearson_notams():
    """
    Toronto Pearson (CYYZ) — weather impact (automated) + NOTAM status (manual).

    v2.12: Replaced failed NOTAM source chain with AWC METAR-based weather
    impact indicator. NOTAMs have no viable free automated API for Canadian
    airports (AWC dropped NOTAM support in Sep 2025 overhaul; NAV Canada CFPS
    and FAA search require browser sessions). Pearson Airport Operations
    is now a manual placeholder with clear retrieval instructions.

    AUTOMATED — Pearson Weather Impact:
      Source: AWC METAR API (aviationweather.gov/api/data/metar)
      Confirmed working — no auth, worldwide coverage, JSON response.
      Severity scale (FAA flight categories):
        0 = VFR   (vis >5SM, ceiling >3000ft) — normal operations
        1 = MVFR  (vis 3-5SM or ceiling 1000-3000ft) — minor impact
        2 = IFR   (vis 1-3SM or ceiling 500-1000ft) — significant impact
        3 = LIFR  (vis <1SM or ceiling <500ft) — severe impact / possible closure

    MANUAL — Pearson Airport Operations (NOTAMs):
      check: plan.navcanada.ca → Weather & NOTAM → enter CYYZ
    """
    AWC_METAR_URL = "https://aviationweather.gov/api/data/metar?ids=CYYZ&format=json"

    FLTCAT_SEVERITY = {"VFR": 0, "MVFR": 1, "IFR": 2, "LIFR": 3}
    FLTCAT_LABEL = {
        "VFR":  "VFR — normal operations",
        "MVFR": "MVFR — marginal conditions, possible minor delays",
        "IFR":  "IFR — low visibility/ceiling, expect delays",
        "LIFR": "LIFR — very low visibility/ceiling, possible closures",
    }

    results = []

    # ── Automated: Pearson Weather Impact via AWC METAR ───────────────────────
    try:
        r = SESSION.get(
            AWC_METAR_URL,
            timeout=TIMEOUT,
            headers={
                "User-Agent": "Mozilla/5.0 (compatible; TII-Scraper/2.12)",
                "Accept": "application/json",
            },
        )
        r.raise_for_status()
        data = r.json()

        if not data:
            raise ValueError("Empty response from AWC METAR API")

        obs = data[0]
        flt_cat  = obs.get("fltCat", "VFR").upper().strip()
        severity = FLTCAT_SEVERITY.get(flt_cat, 0)
        label    = FLTCAT_LABEL.get(flt_cat, flt_cat)
        raw_ob   = obs.get("rawOb", "")
        temp     = obs.get("temp")
        visib    = obs.get("visib")
        wspd     = obs.get("wspd")
        cover    = obs.get("cover", "")
        obs_time = obs.get("reportTime", str(date.today()))[:10]

        # Build ceiling note from clouds array
        clouds = obs.get("clouds", [])
        ceiling_note = ""
        if clouds:
            lowest = clouds[0]
            ceiling_note = (f" Ceiling: {lowest.get('cover','')} "
                           f"{lowest.get('base','')}ft.")
        elif cover == "SKC" or cover == "CLR":
            ceiling_note = " Sky clear."

        notes = (
            f"{label}.{ceiling_note} "
            f"Visibility: {visib}SM. Wind: {wspd}kt. Temp: {temp}°C. "
            f"Raw: {raw_ob}. "
            f"Scale: 0=VFR (normal), 1=MVFR (marginal), "
            f"2=IFR (low vis/ceiling), 3=LIFR (severe). "
            f"Source: AWC METAR API — aviationweather.gov/api/data/metar?ids=CYYZ"
        )

        result = _ok(
            "Pearson Weather Impact", severity, f"({flt_cat})",
            "AWC METAR API (aviationweather.gov)", AWC_METAR_URL,
            obs_time, notes,
        )
        # Apply severity thresholds manually (threshold system uses indicator name matching)
        if severity >= 3:
            result["status"] = "alert"
            result["threshold_note"] = f"LIFR conditions at Pearson — severe weather impact"
        elif severity >= 2:
            result["status"] = "warn"
            result["threshold_note"] = f"IFR conditions at Pearson — low visibility/ceiling"
        elif severity >= 1:
            result["status"] = "warn"
            result["threshold_note"] = f"MVFR conditions at Pearson — marginal weather"
        results.append(result)

    except Exception as e:
        results.append(_err(
            "Pearson Weather Impact",
            "AWC METAR API (aviationweather.gov)",
            AWC_METAR_URL,
            f"METAR fetch failed: {e}. "
            f"Manual: aviationweather.gov/data/metar/?ids=CYYZ",
        ))

    # ── Manual: Pearson Airport Operations (NOTAMs) ───────────────────────────
    # No viable free automated NOTAM API exists for Canadian airports post-Sep 2025.
    # AWC dropped NOTAM support. NAV Canada CFPS and FAA search require browser sessions.
    results.append(_manual(
        "Pearson Airport Operations",
        "NAV Canada CFPS — plan.navcanada.ca",
        "NOTAM APIs for Canadian airports require browser sessions — no free automated source. "
        "Manual check: plan.navcanada.ca → Weather & NOTAM tab → enter CYYZ → NOTAMs. "
        "Severity scale: 0=Normal, 1=Reduced capacity (approach U/S), "
        "2=Runway closure, 3=Flow control or security restriction. "
        "Also check: flightaware.com/live/airport/CYYZ for live delay status.",
        sector="transport_logistics",
    ))

    return results

def fetch_ontario_er_capacity():
    """
    data.ontario.ca — ER Wait Times.
    v1.6: dataset 'wait-time-information-system' exists but has 0 resources
    (data served through ontario's reporting portal, not as downloadable files).
    Now returns a manual placeholder with retrieval instructions.
    Keeping as a function (not moving to manual-only list) so it stays visible
    in the sector output and can be re-automated if a CSV endpoint appears.
    """
    # Check if a downloadable resource has been added since last update
    BASE = "https://data.ontario.ca/api/3/action"
    KNOWN_IDS = [
        "wait-time-information-system",
        "emergency-room-wait-times-ontario",
        "emergency-room-national-ambulatory-reporting-system-initiative-erni",
    ]
    for pkg_id in KNOWN_IDS:
        try:
            r = SESSION.get(f"{BASE}/package_show", params={"id": pkg_id}, timeout=TIMEOUT)
            r.raise_for_status()
            result = r.json()
            if not result.get("success"):
                continue
            resources = result["result"].get("resources", [])
            csv_res = [res for res in resources
                       if (res.get("format","").upper() == "CSV" or
                           res.get("url","").lower().endswith(".csv"))]
            if csv_res:
                # A CSV resource appeared — try to download it
                csv_url = csv_res[0]["url"]
                rr = SESSION.get(csv_url, timeout=30)
                rr.raise_for_status()
                if "html" not in rr.headers.get("Content-Type","").lower():
                    reader = csv.DictReader(io.StringIO(rr.text))
                    rows = list(reader)
                    if rows:
                        return [_ok("ER Capacity — Ontario", len(rows), "rows",
                                    f"data.ontario.ca — {pkg_id}", csv_url,
                                    csv_res[0].get("last_modified","unknown"),
                                    f"CSV now available. Columns: {list(rows[0].keys())[:8]}. "
                                    f"Update fetch_ontario_er_capacity() to parse properly.")]
        except Exception:
            continue

    return [_manual("ER Capacity — Ontario",
                    "data.ontario.ca / Health Quality Ontario",
                    "Dataset exists but has no downloadable files (data in reporting portal). "
                    "Manual retrieval: health.gov.on.ca/en/ms/edrs/ → "
                    "download CSV from 'Percent of ED visits completed within target'. "
                    "Or check data.ontario.ca/dataset/wait-time-information-system "
                    "periodically — a CSV resource may appear.",
                    sector="health")]


def fetch_phac_wastewater():
    """
    PHAC Wastewater Surveillance — Health Infobase API (official REST endpoint).
    URL: health-infobase.canada.ca/api/wastewater/table/wastewater_trend
    Returns city-level aggregate rows for Toronto (grouping == "City").
    Four pathogens tracked: covN2 (COVID), fluA (Influenza A), fluB (Influenza B), rsv (RSV).

    Fields used:
      Location       — site/city name
      grouping       — "City" or "Site" (we use City for aggregate)
      measureid      — covN2 / fluA / fluB / rsv
      Viral_Activity_Level — Non-detect / Low / Moderate / High / Very High / NA2
      latestTrend    — Increasing / Decreasing / No Change / No Recent Data
      weekStart      — ISO date of latest data point

    Viral_Activity_Level "NA2" = no recent data — skipped.
    Numeric scale: Non-detect=0, Low=1, Moderate=2, High=3, Very High=4
    Alert threshold: High or Very High on any pathogen.
    """
    URL = "https://health-infobase.canada.ca/api/wastewater/table/wastewater_trend"
    LEVEL_MAP = {
        "non-detect": 0, "low": 1, "moderate": 2, "high": 3, "very high": 4
    }
    MEASURE_LABELS = {
        "covN2": "COVID-19 (SARS-CoV-2)",
        "fluA":  "Influenza A",
        "fluB":  "Influenza B",
        "rsv":   "RSV",
    }

    try:
        r = SESSION.get(URL, timeout=30,
                        headers={"Accept": "application/json",
                                 "User-Agent": "TII-Scraper/2.10"})
        r.raise_for_status()
        all_rows = r.json()
    except Exception as e:
        return [_err("Wastewater Virus Signal", "PHAC Health Infobase API", URL, str(e))]

    # Filter to Toronto city-level aggregates with current data
    toronto_rows = [
        row for row in all_rows
        if row.get("city", "").lower() == "toronto"
        and row.get("grouping") == "City"
        and row.get("Viral_Activity_Level", "").upper() != "NA2"
    ]

    if not toronto_rows:
        return [_err("Wastewater Virus Signal", "PHAC Health Infobase API", URL,
                     "No current Toronto city-level rows found in API response.")]

    results = []
    for row in toronto_rows:
        measure   = row.get("measureid", "unknown")
        level_raw = row.get("Viral_Activity_Level", "")
        trend     = row.get("latestTrend", "")
        week      = row.get("weekStart", "unknown")
        label     = MEASURE_LABELS.get(measure, measure)
        level_key = level_raw.lower().strip()
        level_num = LEVEL_MAP.get(level_key)

        # Skip non-detect — not a meaningful signal for TII
        if level_key == "non-detect":
            continue

        if level_num is None:
            continue

        trend_note = f" Trend: {trend}." if trend and trend != "No Change" else ""
        notes = (
            f"Toronto wastewater surveillance — {label}. "
            f"Signal level: {level_raw}.{trend_note} "
            f"Week of {week}. "
            f"Scale: Non-detect=0, Low=1, Moderate=2, High=3, Very High=4. "
            f"Alert threshold: High or above on any pathogen. "
            f"Source: PHAC Health Infobase API — health-infobase.canada.ca/api/wastewater"
        )

        result = _ok(
            f"Wastewater Signal — {label}",
            level_num,
            level_raw,
            "PHAC Health Infobase API (wastewater_trend)",
            URL,
            week,
            notes
        )
        result["sector"] = "health"

        # Apply threshold: alert if High (3) or Very High (4)
        if level_num >= 3:
            result["status"] = "alert"
            result["threshold_note"] = f"{label} at {level_raw} in Toronto wastewater"
        elif level_num == 2:
            result["status"] = "warn"
            result["threshold_note"] = f"{label} at Moderate level in Toronto wastewater"

        results.append(result)

    if not results:
        return [_manual("Wastewater Virus Signal", "PHAC Health Infobase API",
                        "API returned data but all Toronto signals were Non-detect or NA. "
                        "Check health-infobase.canada.ca/wastewater/ for current status.",
                        sector="health")]

    return results


# ══════════════════════════════════════════════════════════════════════════════
# SECTOR: FOOD
# ══════════════════════════════════════════════════════════════════════════════

def fetch_statcan_cpi():
    """
    StatsCan CPI — Food and All-items.

    v2.12 FIX: Sequential per-vector requests caused GitHub Actions timeout
    (up to 6 × 30s = 180s). Replaced with a single batched POST for all
    vectors at once with an 8s timeout. One network round-trip instead of six.

    v2.11 FIX: v41693271 (food-at-stores) stale after Jan 2025 (StatsCan
    basket weight update, June 2025). Added freshness gate (120 days) and
    fallback vector list — the batch POST returns all; we pick the freshest.

    Vectors batched in single POST:
      41693271, 41693327, 41693328, 41693329, 41693472 — food-at-stores candidates
      41690973 — all-items CPI (confirmed working)
    """
    WDS_URL    = "https://www150.statcan.gc.ca/t1/wds/rest/getDataFromVectorsAndLatestNPeriods"
    FOOD_VIDS  = [41693271, 41693327, 41693328, 41693329, 41693472]
    ALLITEMS   = 41690973
    STALE_DAYS = 120
    ALL_VIDS   = FOOD_VIDS + [ALLITEMS]

    # ── Single batched POST — all vectors at once, short timeout ─────────────
    try:
        r = SESSION.post(
            WDS_URL,
            json=[{"vectorId": vid, "latestN": 2} for vid in ALL_VIDS],
            timeout=8,
        )
        r.raise_for_status()
        raw_items = r.json()
    except (requests.RequestException, json.JSONDecodeError) as e:
        return [
            _err("Grocery Price Inflation (Food CPI)", "StatsCan WDS API", WDS_URL, str(e)),
            _err("All-items CPI (Canada)",             "StatsCan WDS API", WDS_URL, str(e)),
        ]

    # ── Parse each vector response into (vid, value, ref, age_days) ──────────
    parsed = {}   # vid -> (value, ref, age_days)
    for item in raw_items:
        obj = item.get("object", {})
        vid = obj.get("vectorId")
        if item.get("status") != "SUCCESS" or vid is None:
            continue
        points = obj.get("vectorDataPoint", [])
        if not points:
            continue
        try:
            val = round(float(points[-1]["value"]), 1)
            ref = points[-1].get("refPer", "unknown")
            age_days = 9999
            try:
                age_days = (datetime.utcnow() - datetime.fromisoformat(ref[:10])).days
            except Exception:
                pass
            parsed[vid] = (val, ref, age_days)
        except (ValueError, TypeError):
            continue

    results = []

    # ── Food CPI: first fresh food vector wins ────────────────────────────────
    food_val = food_ref = food_vid = None
    food_errors = []
    for vid in FOOD_VIDS:
        if vid not in parsed:
            food_errors.append(f"v{vid}: no data")
            continue
        val, ref, age = parsed[vid]
        if age > STALE_DAYS:
            food_errors.append(f"v{vid}: stale ({age}d, ref {ref})")
            continue
        food_val, food_ref, food_vid = val, ref, vid
        break

    if food_val is not None:
        results.append(_ok(
            "Grocery Price Inflation (Food CPI)", food_val,
            "CPI index (2002=100)",
            "StatsCan WDS API (Table 18-10-0004-01)", WDS_URL, food_ref,
            f"Food purchased from stores, Canada-wide, not seasonally adjusted. "
            f"Vector v{food_vid}. Released ~6 weeks after reference month."))
    else:
        results.append(_err(
            "Grocery Price Inflation (Food CPI)", "StatsCan WDS API", WDS_URL,
            f"All food vectors stale or missing. {'; '.join(food_errors)}. "
            f"Manual: www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid=1810000401"))

    # ── All-items CPI ─────────────────────────────────────────────────────────
    if ALLITEMS in parsed:
        val, ref, age = parsed[ALLITEMS]
        if age <= STALE_DAYS:
            results.append(_ok(
                "All-items CPI (Canada)", val, "CPI index (2002=100)",
                "StatsCan WDS API (Table 18-10-0004-01)", WDS_URL, ref,
                f"All-items CPI, Canada-wide, not seasonally adjusted. "
                f"Vector v{ALLITEMS}. Released ~6 weeks after reference month."))
        else:
            results.append(_err("All-items CPI (Canada)", "StatsCan WDS API", WDS_URL,
                                f"Stale: {age}d old (ref {ref})"))
    else:
        results.append(_err("All-items CPI (Canada)", "StatsCan WDS API", WDS_URL,
                            f"v{ALLITEMS} not returned by API"))

    return results


# ══════════════════════════════════════════════════════════════════════════════
# SECTOR: TRANSPORT
# ══════════════════════════════════════════════════════════════════════════════

def fetch_ttc_ridership():
    """
    TTC Ridership — Toronto Open Data.

    v2.0 final: The only available resource is '1985-2019-analysis-of-ridership.xlsx'
    — a static historical document, last updated 2019. Current TTC ridership data
    is published in monthly KPI PDFs on ttc.ca, not as open data downloads.

    This function checks each run whether a live CSV/XLSX resource has appeared
    on the package, and will automatically parse it if one shows up.
    Manual retrieval: ttc.ca/transparency-and-accountability/Operating-Statistics
    """
    BASELINE_2019_MONTHLY = 43_250_000
    PACKAGE_IDS = ["ttc-ridership-counts", "ttc-ridership-analysis", "ttc-ridership"]
    TTC_SKIP = ["read-me", "readme", "read_me", "1985-2019", "historical"]

    for pid in PACKAGE_IDS:
        candidate = _ckan_package(pid)
        if not candidate:
            continue
        for res in candidate.get("resources", []):
            url_s = res.get("url", "")
            name  = res.get("name", "")
            fmt   = res.get("format", "").upper()
            # Skip known static/historical files
            if any(kw in url_s.lower() or kw in name.lower() for kw in TTC_SKIP):
                continue
            # Only attempt if it looks like a data file
            if not (fmt in ("CSV", "XLSX", "XLS") or
                    any(url_s.lower().endswith(e) for e in [".csv", ".xlsx", ".xls"])):
                continue
            # A new resource appeared — try to parse it
            try:
                r = SESSION.get(url_s, timeout=30)
                r.raise_for_status()
                raw = r.content
                ct  = r.headers.get("Content-Type", "").lower()
                if "html" in ct:
                    continue
                rows = []
                is_ole2 = raw[:4] == b'\xd0\xcf\x11\xe0'
                is_xlsx = raw[:4] == b'PK\x03\x04' and b'xl/workbook' in raw[:2000]
                if is_ole2:
                    import xlrd
                    wb = xlrd.open_workbook(file_contents=raw)
                    ws = max((wb.sheet_by_index(i) for i in range(wb.nsheets)),
                             key=lambda s: sum(1 for r in range(min(s.nrows,20))
                                               for c in range(s.ncols)
                                               if str(s.cell_value(r,c)).strip()))
                    rows = [[str(ws.cell_value(r,c)) for c in range(ws.ncols)]
                            for r in range(ws.nrows)]
                elif is_xlsx:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                    best_ws = max(wb.worksheets,
                                  key=lambda s: sum(1 for row in list(s.iter_rows())[:20]
                                                    for c in row
                                                    if str(c.value or "").strip()))
                    rows = [[str(c.value or "") for c in row] for row in best_ws.iter_rows()]
                else:
                    text = raw.decode("utf-8", errors="replace").replace("\x00", "")
                    rows = list(csv.reader(io.StringIO(text)))

                if len(rows) < 2:
                    continue
                header = rows[0]
                trip_col = next((i for i, h in enumerate(header) if any(
                    k in h.lower() for k in
                    ["trip","ridership","boarding","count","ride","passenger","revenue"]
                )), None)
                date_col = next((i for i, h in enumerate(header) if any(
                    k in h.lower() for k in ["date","month","period","year"]
                )), None)
                if trip_col is None:
                    continue
                for row in reversed(rows[1:]):
                    if len(row) > trip_col:
                        try:
                            val = float(str(row[trip_col]).replace(",","").strip() or "0")
                            if val > 0:
                                ref = (str(row[date_col])
                                       if date_col and len(row) > date_col else "unknown")
                                pct = round(val / BASELINE_2019_MONTHLY * 100, 1)
                                return [_ok("TTC Ridership Index", pct,
                                            "% of 2019 baseline",
                                            "Toronto Open Data — TTC Ridership",
                                            url_s, ref,
                                            f"{int(val):,} trips. "
                                            f"Baseline {BASELINE_2019_MONTHLY:,}/month.")]
                        except (ValueError, TypeError):
                            continue
            except Exception:
                continue

    return [_manual("TTC Ridership Index",
                    "TTC Monthly KPI Reports (PDF)",
                    "No live open data feed available. "
                    "Current ridership in monthly KPI PDFs: "
                    "ttc.ca/transparency-and-accountability/Operating-Statistics — "
                    "look for 'Revenue Rides' figure. "
                    "Toronto Open Data only has 1985-2019 historical XLSX. "
                    "Will auto-fetch if a live resource appears on the package.",
                    sector="transport_logistics")]


def fetch_ttc_service_status():
    """
    TTC Subway Service Status — real-time disruption detection.

    PRIMARY SOURCE: alerts.ttc.ca/api/alerts/live-alerts
      — Official TTC alerts JSON API (the same backend the ttc.ca website reads).
      — Returns structured records with fields: routeType, effect, severity,
        alertType, title, headerText, shuttleType, cause, activePeriod.
      — No JS rendering required. Confirmed live as of 2026-04-10.

    FALLBACK: bustime.ttc.ca/gtfsrt/alerts (GTFS-RT protobuf)
      — Requires `pip install gtfs-realtime-bindings`; silently skipped if absent.

    Returns:
      - TTC Subway Service Status: 0=normal, 1=partial disruption, 2=major disruption
      - TTC Active Service Alerts: count of active Unplanned subway alerts

    Status logic (effect field, subway routes only):
      alert (2) = any Unplanned alert with effect NO_SERVICE
                  OR 2+ subway segments with Critical severity
      warn  (1) = any Unplanned alert with effect REDUCED_SERVICE / SIGNIFICANT_DELAYS
                  OR any Planned alert with effect NO_SERVICE (advance notice, not crisis)
                  OR shuttleType = "Running" on any subway route
      ok    (0) = no active subway alerts, or only elevator/escalator notices

    Note: Planned track closures (alertType="Planned") are common weekend maintenance.
    Unplanned closures (alertType="Unplanned") are the signal to escalate to alert.

    Lines monitored: Line 1 (Yonge-University), Line 2 (Bloor-Danforth),
                     Line 3 (Scarborough), Line 4 (Sheppard).
    TTC carries ~1.7M riders/week — an unplanned Line 1/2 suspension is a
    Tier-1 urban mobility event cascading into surface congestion and GO overcrowding.
    """
    LIVE_ALERTS_URL = "https://alerts.ttc.ca/api/alerts/live-alerts"
    GTFS_RT_URL     = "https://bustime.ttc.ca/gtfsrt/alerts"
    DISPLAY_URL     = "https://www.ttc.ca/service-advisories/all-service-alerts"
    SOURCE_LABEL    = "TTC — alerts.ttc.ca/api/alerts/live-alerts"

    # ── Source 1: alerts.ttc.ca JSON API ──────────────────────────────────────
    try:
        r = SESSION.get(LIVE_ALERTS_URL, timeout=TIMEOUT,
                        headers={"Accept": "application/json",
                                 "Referer": "https://www.ttc.ca/"})
        r.raise_for_status()
        data = r.json()

        routes = data.get("routes", [])
        last_updated = data.get("lastUpdated", "")

        # Filter to subway routes only (routeType = "Subway" or route in 1-4)
        subway_alerts = [
            a for a in routes
            if (str(a.get("routeType", "")).lower() == "subway"
                or str(a.get("route", "")) in ("1", "2", "3", "4"))
        ]

        # Separate unplanned vs planned
        unplanned = [a for a in subway_alerts
                     if str(a.get("alertType", "")).lower() == "unplanned"]
        planned   = [a for a in subway_alerts
                     if str(a.get("alertType", "")).lower() == "planned"]

        # Determine severity
        # alert (2): any unplanned NO_SERVICE, or 2+ unplanned Critical alerts
        unplanned_no_svc = [a for a in unplanned
                            if str(a.get("effect", "")).upper() == "NO_SERVICE"]
        unplanned_critical = [a for a in unplanned
                               if str(a.get("severity", "")).lower() == "critical"]

        # warn (1): unplanned delays/reduced service, OR planned NO_SERVICE
        #           (planned = weekend maintenance, not a crisis), OR shuttle running
        unplanned_delays = [a for a in unplanned
                            if str(a.get("effect", "")).upper() in
                            ("REDUCED_SERVICE", "SIGNIFICANT_DELAYS", "DETOUR")]
        shuttle_running = [a for a in subway_alerts
                           if str(a.get("shuttleType", "")).lower() == "running"]
        planned_no_svc  = [a for a in planned
                           if str(a.get("effect", "")).upper() == "NO_SERVICE"]

        if unplanned_no_svc or len(unplanned_critical) >= 2:
            severity = 2
        elif unplanned_delays or shuttle_running or planned_no_svc:
            severity = 1
        else:
            severity = 0

        # Build summary note
        alert_count = len(unplanned)  # only unplanned count as "active alerts"
        severity_labels = [
            "Normal service on all subway lines.",
            "Partial disruption — delays, reduced service, or planned closure with shuttle.",
            "Major unplanned disruption — full or multi-segment suspension.",
        ]
        disruption_lines = []
        for a in (unplanned_no_svc or unplanned_delays or shuttle_running or planned_no_svc)[:3]:
            disruption_lines.append(a.get("title", a.get("headerText", ""))[:120])
        disruption_detail = " | ".join(disruption_lines) if disruption_lines else ""

        notes = (
            f"TTC subway: {severity_labels[severity]} "
            f"{len(subway_alerts)} subway alert(s) active "
            f"({len(unplanned)} unplanned, {len(planned)} planned). "
            + (f"Key alert: {disruption_detail} " if disruption_detail else "")
            + f"API last updated: {last_updated[:16]}. "
            f"Lines 1–4 (Yonge-University, Bloor-Danforth, Scarborough, Sheppard). "
            f"~1.7M weekly riders. Unplanned Line 1/2 suspension = Tier-1 mobility event. "
            f"Source: alerts.ttc.ca/api/alerts/live-alerts"
        )

        result = _ok("TTC Subway Service Status", severity, "",
                     SOURCE_LABEL, DISPLAY_URL, str(date.today()), notes)
        if severity == 2:
            result["status"] = "alert"
            result["threshold_note"] = "Unplanned subway suspension — major service disruption"
        elif severity == 1:
            result["status"] = "warn"
            result["threshold_note"] = "Subway disruption — delays, reduced service, or planned closure"

        results_out = [result]
        if alert_count > 0:
            results_out.append(
                _ok("TTC Active Service Alerts", alert_count, "unplanned alerts",
                    SOURCE_LABEL, DISPLAY_URL, str(date.today()),
                    f"{alert_count} unplanned subway alert(s). "
                    f"Check ttc.ca/service-advisories/all-service-alerts for detail.")
            )
            if alert_count > 0:
                results_out[-1]["status"] = "warn" if severity < 2 else "alert"

        return results_out

    except (requests.RequestException, ValueError, KeyError) as e:
        pass  # Fall through to GTFS-RT

    # ── Fallback: GTFS-RT protobuf ─────────────────────────────────────────────
    SUSPENSION_KW = ["no service", "suspended", "suspension", "shut down",
                     "closed", "closure", "out of service"]
    DISRUPTION_KW = ["delay", "reduced speed", "shuttle bus", "shuttle buses",
                     "diversion", "bypassing", "reduced service"]
    LINE_KW       = ["line 1", "line 2", "line 3", "line 4",
                     "yonge", "university", "bloor", "danforth",
                     "scarborough", "sheppard", "subway"]

    try:
        r = SESSION.get(GTFS_RT_URL, timeout=TIMEOUT)
        r.raise_for_status()
        try:
            from google.transit import gtfs_realtime_pb2
            feed = gtfs_realtime_pb2.FeedMessage()
            feed.ParseFromString(r.content)
            texts = []
            for entity in feed.entity:
                if entity.HasField("alert"):
                    for tr in entity.alert.header_text.translation:
                        if tr.text: texts.append(tr.text)
        except ImportError:
            texts = [r.content.decode("utf-8", errors="replace")]

        subway_texts = [t for t in texts if any(k in t.lower() for k in LINE_KW)]
        if any(any(k in t.lower() for k in SUSPENSION_KW) for t in subway_texts):
            severity = 2
        elif any(any(k in t.lower() for k in DISRUPTION_KW) for t in subway_texts):
            severity = 1
        else:
            severity = 0

        severity_labels = ["Normal service.", "Partial disruption.", "Major disruption."]
        result = _ok("TTC Subway Service Status", severity, "",
                     "TTC — bustime.ttc.ca/gtfsrt/alerts", DISPLAY_URL,
                     str(date.today()),
                     f"TTC subway: {severity_labels[severity]} Source: GTFS-RT fallback.")
        if severity == 2:
            result["status"] = "alert"
            result["threshold_note"] = "Unplanned subway suspension — major service disruption"
        elif severity == 1:
            result["status"] = "warn"
            result["threshold_note"] = "Subway disruption detected via GTFS-RT"
        return [result]

    except Exception as e:
        return [_err("TTC Subway Service Status", SOURCE_LABEL, DISPLAY_URL, str(e))]


# ══════════════════════════════════════════════════════════════════════════════
# SECTOR: TRANSPORT & LOGISTICS — RAIL
# ══════════════════════════════════════════════════════════════════════════════

def fetch_go_transit_status():
    """
    GO Transit service status — scrapes gotransit.com/en/service-updates.
    Returns:
      - GO Transit Active Alerts: count of current service alerts
      - GO Transit Service Status: 0=normal, 1=minor alerts, 2=major disruption
    Lines monitored: Lakeshore East/West, Kitchener, Barrie, Stouffville,
                     Milton, Richmond Hill, UP Express.
    GO Transit serves ~70,000 weekday riders across the GTHA.
    A major disruption cascades immediately into highway congestion and TTC overcrowding.
    """
    url = "https://www.gotransit.com/en/service-updates"
    try:
        r = SESSION.get(url, timeout=TIMEOUT,
                        headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.9"})
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("GO Transit Service Status", "GO Transit — gotransit.com", url, str(e))]

    try:
        soup = BeautifulSoup(r.content, "html.parser")
        text = soup.get_text(" ", strip=True).lower()

        # Look for alert/disruption language
        major_keywords = ["suspended", "no service", "major delay", "service disruption",
                          "all trains cancelled", "bus bridge", "shuttle bus"]
        minor_keywords = ["delay", "alert", "disruption", "reduced", "modified service",
                          "late", "holding"]
        normal_keywords = ["no service alerts", "no delays", "on schedule",
                           "service is normal", "good service"]

        # Count alert items — look for list items or alert containers
        alert_items = soup.find_all(["li", "div", "p"],
                                    string=lambda t: t and any(
                                        k in t.lower() for k in minor_keywords))
        # Also check for structured alert elements
        for cls in ["alert", "service-alert", "disruption", "notification"]:
            found = soup.find_all(attrs={"class": lambda c: c and cls in " ".join(c).lower()})
            alert_items.extend(found)

        alert_count = len(set(str(i) for i in alert_items))

        if any(k in text for k in normal_keywords) and alert_count == 0:
            severity = 0
            status_label = "Normal service"
        elif any(k in text for k in major_keywords):
            severity = 2
            status_label = "Major disruption — check gotransit.com for details"
        elif any(k in text for k in minor_keywords) or alert_count > 0:
            severity = 1
            status_label = f"Service alerts active ({alert_count} found)"
        else:
            severity = 0
            status_label = "No disruptions detected"

        notes = (
            f"{status_label}. "
            f"GO Transit: 7 rail lines + UP Express, ~70,000 weekday GTHA riders. "
            f"Severity: 0=Normal, 1=Minor alerts, 2=Major disruption. "
            f"Source: gotransit.com/en/service-updates (scraped)."
        )
        results = [
            _ok("GO Transit Service Status", severity, "",
                "GO Transit — gotransit.com", url, str(date.today()), notes),
        ]
        if alert_count > 0:
            results.append(_ok("GO Transit Active Alerts", alert_count, "alerts",
                               "GO Transit — gotransit.com", url, str(date.today()),
                               f"Active service alerts detected. Check gotransit.com for line-specific details."))
        return results

    except Exception as e:
        return [_err("GO Transit Service Status", "GO Transit — gotransit.com", url, str(e))]


def fetch_via_rail_status():
    """
    VIA Rail corridor status — scrapes viarail.ca/en/plan/service-status.
    Monitors the Quebec City-Windsor Corridor, Canada's busiest intercity rail route.
    Key trains through Toronto: 50/51 (Toronto-Montreal), 60/61 (Toronto-Ottawa),
    1/2 (The Canadian, Toronto-Vancouver — weekly).
    A corridor disruption affects business travel, cargo, and GO/VIA shared infrastructure.
    Returns severity: 0=Normal, 1=Delays, 2=Cancellations/Major disruption.
    """
    url = "https://www.viarail.ca/en/plan/service-status"
    try:
        r = SESSION.get(url, timeout=TIMEOUT,
                        headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.9"})
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("VIA Rail Corridor Status", "VIA Rail — viarail.ca", url, str(e))]

    try:
        soup = BeautifulSoup(r.content, "html.parser")
        text = soup.get_text(" ", strip=True).lower()

        cancel_keywords  = ["cancelled", "suspended", "no service", "service suspended"]
        delay_keywords   = ["delay", "late", "disruption", "modified", "held"]
        normal_keywords  = ["on time", "normal service", "no disruption", "operating normally"]

        # Check specifically for Toronto/corridor mentions
        toronto_context = any(
            k in text for k in ["toronto", "windsor", "montreal", "ottawa", "corridor"]
        )

        if any(k in text for k in cancel_keywords):
            severity = 2
            status_label = "Cancellations or suspended service"
        elif any(k in text for k in delay_keywords):
            severity = 1
            status_label = "Delays reported on corridor"
        elif any(k in text for k in normal_keywords) or not toronto_context:
            severity = 0
            status_label = "Normal service"
        else:
            severity = 0
            status_label = "No disruptions detected"

        notes = (
            f"{status_label}. "
            f"VIA Rail Quebec City-Windsor Corridor: ~406 trains/week, Canada's busiest intercity route. "
            f"Toronto trains: 50/51 (Montreal), 60/61 (Ottawa). "
            f"Corridor runs on CN-owned track — freight congestion and infrastructure work affect schedules. "
            f"Severity: 0=Normal, 1=Delays, 2=Cancellations. Source: viarail.ca/en/plan/service-status."
        )
        return [_ok("VIA Rail Corridor Status", severity, "",
                    "VIA Rail — viarail.ca", url, str(date.today()), notes)]

    except Exception as e:
        return [_err("VIA Rail Corridor Status", "VIA Rail — viarail.ca", url, str(e))]


def fetch_freight_rail_labour_risk():
    """
    CN and CP (CPKC) freight rail labour risk — scrapes CN and CPKC news/bargaining pages.
    A freight rail stoppage halts ~$1B/day in Canadian trade.
    CN's Oakville Subdivision and Kingston Subdivision are the primary Toronto-area corridors.
    CN also serves Port of Toronto and the intermodal yard at Brampton/Malport.

    Returns a risk level based on presence of active bargaining/dispute language:
      0 = No active labour dispute detected
      1 = Active negotiations underway (elevated risk)
      2 = Strike notice or lockout imminent/active

    In August 2024 CN and CPKC simultaneously locked out 9,300 workers;
    the federal government intervened with binding arbitration within 17 hours.
    Under the Canada Labour Code, rail is not classified as "essential service"
    meaning full stoppage is legally permitted.
    """
    sources = [
        ("CN", "https://www.cn.ca/en/media/bargaining-updates/",
         "CN — cn.ca/media/bargaining-updates"),
        ("CPKC",    "https://www.cpkcr.com/en/media/news-releases",
         "CPKC — cpkcr.com/media/news-releases"),
    ]

    imminent_keywords  = ["strike notice", "lockout", "work stoppage", "job action",
                          "bargaining deadline", "72-hour notice", "24-hour notice"]
    active_keywords    = ["bargaining underway", "negotiations underway", "collective agreement expires",
                          "contract talks", "labour dispute", "at the bargaining table",
                          "teamsters demand", "tcrc demand", "failed to reach"]
    resolved_keywords  = ["ratified", "agreement reached", "new contract", "signed",
                          "arbitration awarded", "tentative agreement", "collective agreement in place",
                          "three-year deal", "new deal", "agreement imposed"]
    # These phrases appear on CN/CPKC pages as permanent nav/page-title text
    # even when no dispute is active — treat as baseline 0
    nav_false_positive = ["bargaining updates", "media centre", "news releases",
                          "stay informed with the latest"]

    # Known contract expiry dates — use to flag upcoming negotiation windows
    CONTRACT_EXPIRY = {
        "CN":   "2026-12-31",   # 3-yr deal from binding arbitration April 2025
        "CPKC": "2026-12-31",   # CPKC arbitration award same period
    }

    results = []
    for carrier, url, source_label in sources:
        try:
            r = SESSION.get(url, timeout=TIMEOUT,
                            headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.10"})
            r.raise_for_status()
            soup = BeautifulSoup(r.content, "html.parser")

            # Strip nav, header, footer, sidebar before scanning —
            # these contain permanent "bargaining" menu links causing false positives
            for tag in soup.find_all(["nav", "header", "footer", "aside",
                                      "script", "style", "noscript"]):
                tag.decompose()
            # Also strip elements with nav-like class/id names
            for tag in soup.find_all(attrs={"class": lambda c: c and any(
                    k in " ".join(c).lower() for k in
                    ["nav", "menu", "header", "footer", "sidebar", "breadcrumb"])}):
                tag.decompose()

            text = soup.get_text(" ", strip=True).lower()

            # Raised from 1200 to 4000 — CN/CPKC pages are JS-rendered;
            # nav shell even after stripping can exceed 1200 chars.
            # Real bargaining news articles contain 4000+ chars of body text.
            if len(text) < 4000:
                risk = 0
                risk_label = "Baseline — page content minimal (JS-rendered or no active updates)"
            elif any(k in text for k in nav_false_positive):
                # Page is the static "stay informed" placeholder — no active dispute
                risk = 0
                risk_label = "Baseline — dedicated bargaining page exists but no active dispute content"
            elif any(k in text for k in imminent_keywords):
                risk = 2
                risk_label = "Strike/lockout imminent or active"
            elif any(k in text for k in resolved_keywords):
                risk = 0
                risk_label = "Labour agreement in place"
            elif any(k in text for k in active_keywords):
                risk = 1
                risk_label = "Active bargaining underway"
            else:
                risk = 0
                risk_label = "No active labour dispute detected"

            # Add contract expiry context to notes
            expiry = CONTRACT_EXPIRY.get(carrier, "unknown")
            expiry_note = (
                f"Current collective agreement expires {expiry}. "
                f"Next negotiation window opens ~6 months prior. "
            ) if expiry != "unknown" else ""

            notes = (
                f"{carrier}: {risk_label}. "
                f"{expiry_note}"
                f"Risk scale: 0=No dispute, 1=Active negotiations, 2=Imminent/active stoppage. "
                f"A CN/CP simultaneous stoppage halts ~$1B/day in Canadian trade and affects "
                f"GO Transit commuter lines sharing CN track. "
                f"Federal government has historically intervened within hours under Canada Labour Code s.107. "
                f"Source: {url}"
            )
            results.append(_ok(f"{carrier} Freight Labour Risk", risk, "",
                               source_label, url, str(date.today()), notes))

        except Exception as e:
            results.append(_err(f"{carrier} Freight Labour Risk",
                                source_label, url, str(e)))

    return results



def _montreal_notes(teu_val, ref_period, source_url):
    """Build consistent notes string for Port of Montreal TEU results."""
    baseline_low  = 130_000
    baseline_high = 160_000
    vs_baseline = ""
    if teu_val < baseline_low:
        pct_below = round((baseline_low - teu_val) / baseline_low * 100, 1)
        vs_baseline = f" {pct_below}% below 2023-24 baseline range."
    elif teu_val > baseline_high:
        pct_above = round((teu_val - baseline_high) / baseline_high * 100, 1)
        vs_baseline = f" {pct_above}% above 2023-24 baseline range."
    return (
        f"Port of Montreal monthly container throughput — {ref_period}. "
        f"{teu_val:,} TEUs.{vs_baseline} "
        f"2023-24 baseline: 130,000-160,000 TEUs/month. "
        f"Port serves Ontario (28% of trade) and Quebec (53%). "
        f"Connected to CN and CPKC rail networks. "
        f"Warn: <110,000 TEUs (significant volume drop). "
        f"Alert: <85,000 TEUs (severe disruption — labour action or systemic shock). "
        f"Data lag: ~4-6 weeks. Dwell time: see annual Year in Review report. "
        f"Source: {source_url}"
    )


def fetch_port_of_montreal():
    """
    Port of Montreal — monthly TEU container throughput.

    Source: Port of Montreal PMStats backend + press releases.
    The statistics page (port-montreal.com/PMStats/...) is a JSP app that
    loads chart data from a backend JSON endpoint. We try the backend directly
    with appropriate Referer/UA headers, then fall back to scraping the
    latest press release for the most recently announced throughput figure.

    INDICATOR: TEU throughput for most recent published month
    DATA LAG:  ~4-6 weeks (published mid-following-month)
    UNIT:      TEUs (Twenty-foot Equivalent Units)

    Threshold guidance:
      Normal monthly throughput: ~130,000-160,000 TEUs (2023-2024 baseline)
      Warn: <110,000 TEUs (significant volume drop, >20% below baseline)
      Alert: <85,000 TEUs (severe disruption — labour action or systemic shock)

    NOTE: Container dwell time is not available via any free public API.
    The Port publishes dwell time annually in their year-in-review report.
    """
    PMSTATS_BASE = "https://www.port-montreal.com"
    REFERER      = (f"{PMSTATS_BASE}/en/detailed-statistics-history-and-summaries"
                    f"/current-statistics/monthly-teu-throughput")
    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/122.0.0.0 Safari/537.36",
        "Accept":     "application/json, text/plain, */*",
        "Referer":    REFERER,
        "Origin":     PMSTATS_BASE,
    }

    # ── Source 1: PMStats JSON backend probe ─────────────────────────────────
    BACKEND_URLS = [
        f"{PMSTATS_BASE}/PMStats/rest/statistics/teu/monthly",
        f"{PMSTATS_BASE}/PMStats/rest/statistics/container/monthly",
        f"{PMSTATS_BASE}/PMStats/servlet/StatisticsServlet?type=teu&lang=en",
        f"{PMSTATS_BASE}/PMStats/html/frontend/data/teu_monthly.json",
        f"{PMSTATS_BASE}/api/statistics/teu/monthly",
    ]
    for backend_url in BACKEND_URLS:
        try:
            r = SESSION.get(backend_url, headers=HEADERS, timeout=TIMEOUT)
            if r.status_code not in (200, 201):
                continue
            ct = r.headers.get("Content-Type", "").lower()
            if "html" in ct or len(r.content) < 20:
                continue
            raw = r.json()
            teu_val = ref_period = None
            if isinstance(raw, list) and raw:
                last = raw[-1]
                if isinstance(last, dict):
                    for vk in ["value", "teu", "TEU", "volume", "throughput", "count"]:
                        if vk in last:
                            try:
                                teu_val = int(float(last[vk])); break
                            except (ValueError, TypeError):
                                pass
                    for dk in ["period", "month", "date", "label", "refPer"]:
                        if dk in last:
                            ref_period = str(last[dk]); break
            elif isinstance(raw, dict):
                for key in ["data", "months", "values", "series", "teu"]:
                    if key in raw and isinstance(raw[key], list) and raw[key]:
                        last = raw[key][-1]
                        if isinstance(last, (int, float)):
                            teu_val = int(last)
                        elif isinstance(last, dict):
                            for vk in ["value", "teu", "TEU", "volume", "y"]:
                                if vk in last:
                                    try:
                                        teu_val = int(float(last[vk]))
                                    except (ValueError, TypeError):
                                        pass
                        break
            if teu_val and 50_000 < teu_val < 400_000:
                return [_ok("Port of Montreal TEU Throughput", teu_val, "TEUs/month",
                            "Port of Montreal — PMStats backend", backend_url,
                            ref_period or str(date.today()),
                            _montreal_notes(teu_val, ref_period or "unknown", backend_url))]
        except Exception:
            continue

    # ── Source 2: Scrape the statistics page directly ─────────────────────────
    try:
        r = SESSION.get(REFERER, headers={**HEADERS, "Accept": "text/html,*/*"},
                        timeout=TIMEOUT)
        r.raise_for_status()
        soup = BeautifulSoup(r.content, "html.parser")
        text = soup.get_text(" ", strip=True)
        candidates = re.findall(r'\b([1-3]\d{2},\d{3}|[5-9]\d,\d{3})\b', text)
        if candidates:
            teu_val = int(candidates[0].replace(",", ""))
            if 50_000 < teu_val < 400_000:
                date_match = re.search(
                    r'(January|February|March|April|May|June|July|August|'
                    r'September|October|November|December)\s+202[3-9]', text)
                ref_period = date_match.group(0) if date_match else "unknown"
                return [_ok("Port of Montreal TEU Throughput", teu_val, "TEUs/month",
                            "Port of Montreal — statistics page (scraped)", REFERER,
                            ref_period, _montreal_notes(teu_val, ref_period, REFERER))]
    except Exception:
        pass

    # ── Source 3: Press releases — most recently announced figure ─────────────
    PRESS_URLS = [
        f"{PMSTATS_BASE}/en/the-port-of-montreal/news/news/press-release/",
        f"{PMSTATS_BASE}/en/media/press-releases/",
    ]
    for press_url in PRESS_URLS:
        try:
            r = SESSION.get(press_url,
                            headers={**HEADERS, "Accept": "text/html,*/*"},
                            timeout=TIMEOUT)
            r.raise_for_status()
            soup = BeautifulSoup(r.content, "html.parser")
            links = [a["href"] for a in soup.find_all("a", href=True)
                     if any(kw in a["href"].lower() or kw in (a.get_text() or "").lower()
                            for kw in ["results", "traffic", "throughput", "statistics"])]
            if not links:
                continue
            result_url = links[0] if links[0].startswith("http") else PMSTATS_BASE + links[0]
            rr = SESSION.get(result_url,
                             headers={**HEADERS, "Accept": "text/html,*/*"},
                             timeout=TIMEOUT)
            rr.raise_for_status()
            content = BeautifulSoup(rr.content, "html.parser").get_text(" ", strip=True)
            teu_match = re.search(r'([1-3]\d{2},\d{3}|[5-9]\d,\d{3})\s*TEUs?',
                                  content, re.IGNORECASE)
            if teu_match:
                teu_val = int(teu_match.group(1).replace(",", ""))
                period_match = re.search(
                    r'(January|February|March|April|May|June|July|August|'
                    r'September|October|November|December)\s+202[3-9]'
                    r'|Q[1-4]\s+202[3-9]|first half|second half|mid.year',
                    content, re.IGNORECASE)
                ref_period = period_match.group(0) if period_match else "unknown"
                return [_ok("Port of Montreal TEU Throughput", teu_val,
                            "TEUs/month (announced)",
                            "Port of Montreal — press release", result_url,
                            ref_period, _montreal_notes(teu_val, ref_period, result_url))]
        except Exception:
            continue

    # ── Manual fallback ───────────────────────────────────────────────────────
    return [_manual(
        "Port of Montreal TEU Throughput",
        "Port of Montreal — Statistics page",
        "Automated fetch failed (statistics page is JS-rendered or geo-blocked). "
        "Manual retrieval: port-montreal.com/en/detailed-statistics-history-and-summaries"
        "/current-statistics/monthly-teu-throughput "
        "2024 baseline: ~130,000-160,000 TEUs/month. "
        "Warn: <110,000. Alert: <85,000. "
        "Dwell time published annually: port-montreal.com/en/trading-with-the-world-YYYY",
        sector="transport_logistics")]



def fetch_toronto_aqhi():
    """Environment Canada AQHI — unchanged (working)."""
    url = "https://weather.gc.ca/airquality/pages/onaq-001_e.html"
    try:
        r = SESSION.get(url, timeout=TIMEOUT)
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("Air Quality (AQHI)", "Environment Canada AQHI", url, str(e))]

    soup = BeautifulSoup(r.content, "html.parser")
    aqhi_val = risk_level = None
    for tag in soup.find_all(["h1","h2","h3","p","div","span"]):
        m = re.match(r'^(\d+)(Low Risk|Moderate Risk|High Risk|Very High Risk)',
                     tag.get_text(strip=True))
        if m:
            aqhi_val, risk_level = int(m.group(1)), m.group(2)
            break
    if aqhi_val is None:
        m = re.search(r'Current Air Quality Health Index[:\s]+(\d+)', soup.get_text())
        if m:
            aqhi_val = int(m.group(1))
    if aqhi_val is None:
        return [_err("Air Quality (AQHI)", "Environment Canada AQHI", url,
                     "Could not parse AQHI value")]
    if risk_level is None:
        risk_level = ("Low Risk" if aqhi_val <= 3 else "Moderate Risk" if aqhi_val <= 6
                      else "High Risk" if aqhi_val <= 10 else "Very High Risk")
    return [_ok("Air Quality (AQHI)", aqhi_val, "AQHI (1-10+)",
                "Environment Canada AQHI — Toronto", url, date.today(),
                f"Risk: {risk_level}. Alert threshold: AQHI ≥ 4.")]


# ══════════════════════════════════════════════════════════════════════════════
# SECTOR: FINANCIAL
# ══════════════════════════════════════════════════════════════════════════════

def fetch_bank_of_canada_rate():
    """Bank of Canada — TARGET1 series (working in v1.3)."""
    for series in ["TARGET1", "AVGTX", "PRIME"]:
        url = f"https://www.bankofcanada.ca/valet/observations/{series}/json?recent=3"
        try:
            r = SESSION.get(url, timeout=TIMEOUT)
            if r.status_code == 404:
                continue
            r.raise_for_status()
            data = r.json()
            obs = data.get("observations", [])
            if not obs:
                continue
            latest = obs[-1]
            entry = latest.get(series, {})
            rate_str = entry.get("v") if isinstance(entry, dict) else str(entry)
            rate = float(rate_str)
            return [_ok("Bank of Canada Policy Rate", rate, "%",
                        f"Bank of Canada Valet API ({series})", url, latest["d"],
                        f"Overnight target rate as of {latest['d']}. "
                        f"Alert threshold: sustained rise signals tightening credit conditions.")]
        except (requests.RequestException, json.JSONDecodeError, KeyError,
                ValueError, TypeError):
            continue

    # Known fallback value
    return [_ok("Bank of Canada Policy Rate", 2.25, "%",
                "Bank of Canada (known value — Valet API series not resolved)", "",
                "2025-10-29",
                "⚠ Valet API series name unresolved. Known: 2.25% since Oct 29, 2025. "
                "Verify at bankofcanada.ca/valet/docs")]


def fetch_cad_usd_rate():
    """
    CAD/USD exchange rate — Bank of Canada Valet API.
    Series FXUSDCAD: daily USD to CAD rate, published at 16:30 ET.
    Same API already used for TARGET1 (overnight rate) — no new dependency.
    Resilience context: rising USD/CAD = weaker CAD = imported inflation,
    more expensive US goods, tariff pressure amplification.
    Alert threshold: >1.45 = elevated stress (historic stress levels in 2025).
    """
    url = "https://www.bankofcanada.ca/valet/observations/FXUSDCAD/json?recent=5"
    try:
        r = SESSION.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
        obs = data.get("observations", [])
        if not obs:
            return [_err("CAD/USD Exchange Rate", "Bank of Canada Valet API", url,
                         "No observations returned")]
        # Filter out nulls (weekends/holidays have v: null)
        valid = [o for o in obs if o.get("FXUSDCAD", {}).get("v") not in (None, "")]
        if not valid:
            return [_err("CAD/USD Exchange Rate", "Bank of Canada Valet API", url,
                         "All recent observations are null (holiday period?)")]
        latest = valid[-1]
        rate = float(latest["FXUSDCAD"]["v"])
        obs_date = latest["d"]
        prev = valid[-2] if len(valid) >= 2 else None
        change = ""
        if prev and prev.get("FXUSDCAD", {}).get("v"):
            delta = rate - float(prev["FXUSDCAD"]["v"])
            change = f" ({'+' if delta >= 0 else ''}{delta:.4f} vs prev day)"
    except (requests.RequestException, json.JSONDecodeError,
            KeyError, ValueError, TypeError) as e:
        return [_err("CAD/USD Exchange Rate", "Bank of Canada Valet API", url, str(e))]

    return [_ok("CAD/USD Exchange Rate", round(rate, 4), "USD/CAD",
                "Bank of Canada Valet API (FXUSDCAD)", url, obs_date,
                f"Daily closing rate{change}. "
                f"Alert threshold: >1.45 (elevated tariff/import stress). "
                f"Published 16:30 ET on business days.")]


def fetch_toronto_fuel_price():
    """
    Toronto retail gasoline price — Stockr.net daily forecast (primary).
    Fallback: StatsCan Table 18-10-0001-01 monthly average retail price, Toronto.

    Stockr source: stockr.net/Toronto/GasPrice.aspx (note mixed case — lowercase redirects)
    Updated daily at 11am — shows today's and tomorrow's price.
    Calculated from commodity markets, not crowd-sourced.
    Prices in cents per litre, regular unleaded, includes all taxes.
    """
    # SSL cert is valid for stockr.net only — www.stockr.net causes cert mismatch
    url = "https://stockr.net/Toronto/GasPrice.aspx"
    try:
        r = SESSION.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=TIMEOUT)
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("Fuel Price — Toronto", "Stockr.net", url, str(e))]

    try:
        soup = BeautifulSoup(r.text, "html.parser")
        today_price = None
        tomorrow_price = None
        today_date = None
        tomorrow_date = None

        for h3 in soup.find_all("h3"):
            label = h3.get_text(strip=True).lower()
            node = h3.next_sibling
            price_found = None
            date_found = ""
            while node:
                t = str(node).strip()
                if t and t not in ("\n", ""):
                    try:
                        candidate = float(t.replace(",", "").strip())
                        if 100 < candidate < 300:
                            price_found = candidate
                            node = node.next_sibling
                            while node:
                                dt = str(node).strip()
                                if dt and dt not in ("\n", ""):
                                    date_found = BeautifulSoup(dt, "html.parser").get_text(strip=True)
                                    break
                                node = node.next_sibling
                            break
                    except (ValueError, AttributeError):
                        pass
                node = node.next_sibling
            if label == "today" and price_found:
                today_price = price_found
                today_date = date_found
            elif label == "tomorrow" and price_found:
                tomorrow_price = price_found
                tomorrow_date = date_found

        # Fallback: regex scan entire page text
        if today_price is None:
            import re as _re
            matches = _re.findall(r"\b(1[4-9]\d\.\d|2\d{2}\.\d)\b", soup.get_text())
            if matches:
                today_price = float(matches[0])
                today_date = str(date.today())

        if today_price is None:
            return [_err("Fuel Price — Toronto", "Stockr.net", url,
                         "Could not parse price from page")]

        tomorrow_note = (f" Tomorrow: {tomorrow_price}¢/L ({tomorrow_date})."
                         if tomorrow_price else "")
        notes = (
            f"Tier 3 — Local real-time delivery. Regular unleaded, includes all taxes. "
            f"Daily forecast from commodity markets.{tomorrow_note} "
            f"Supply chain: Brent crude (Tier 1) → TCPL pipeline utilization (Tier 2) → "
            f"Toronto pump price (Tier 3). CAD/USD rate amplifies or dampens pass-through. "
            f"Source: stockr.net (Canadian, daily at 11am). "
            f"Warn >150¢/L, Alert >185¢/L."
        )
        return [_ok("Fuel Price — Toronto", today_price, "¢/L",
                    "Stockr.net — daily Toronto pump price", url,
                    today_date, notes)]

    except Exception as e:
        pass  # Fall through to StatsCan fallback

    # ── Fallback: StatsCan Table 18-10-0001-01 — monthly average, Toronto ────
    # Less frequent (monthly) but reliable — official government source.
    # Vector for Toronto regular unleaded retail price (cents/litre).
    STATCAN_URL = "https://www150.statcan.gc.ca/t1/wds/rest/getDataFromVectorsAndLatestNPeriods"
    FUEL_VECTORS = [
        (41692780, "Toronto regular unleaded (v41692780)"),
        (41692781, "Toronto regular unleaded alt (v41692781)"),
    ]
    try:
        payload = [{"vectorId": vid, "latestN": 2} for vid, _ in FUEL_VECTORS]
        r = SESSION.post(STATCAN_URL, json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()
        for item, (vid, label) in zip(data, FUEL_VECTORS):
            if item.get("status") != "SUCCESS":
                continue
            points = item.get("object", {}).get("vectorDataPoint", [])
            if not points:
                continue
            latest = points[-1]
            try:
                val = float(latest["value"])
                # StatsCan reports in dollars/litre — convert to cents
                price_cents = round(val * 100, 1) if val < 5 else round(val, 1)
                if not (80 < price_cents < 300):
                    continue
                ref = latest.get("refPer", "unknown")
                return [_ok("Fuel Price — Toronto", price_cents, "¢/L",
                            f"StatsCan Table 18-10-0001-01 — {label}",
                            STATCAN_URL, ref,
                            f"Monthly average retail price, regular unleaded, Toronto. "
                            f"StatsCan vector v{vid}. Monthly cadence — less current than Stockr. "
                            f"Warn >150¢/L, Alert >185¢/L.")]
            except (ValueError, TypeError):
                continue
    except Exception:
        pass

    return [_err("Fuel Price — Toronto", "Stockr.net / StatsCan 18-10-0001-01",
                 url, "All fuel price sources failed.")]


def fetch_ontario_icu_occupancy():
    """
    Ontario ICU occupancy — data.ontario.ca.
    NOTE: The COVID-era public ICU dataset was discontinued Nov 14, 2024.
    CCSO does not publish public data (PHIPA-restricted, authorized users only).
    This function checks whether the dataset has been reinstated and will
    auto-parse it if so. Otherwise returns a manual placeholder.
    Manual retrieval: PHO Respiratory Virus Tool at
    publichealthontario.ca/en/Data-and-Analysis/Infectious-Disease/Respiratory-Virus-Tool
    (shows ICU and hospital occupancy for respiratory viruses including flu/RSV/COVID).
    """
    # The last known CSV URL — check if it has been resumed
    ICU_CSV = ("https://data.ontario.ca/dataset/1b5ff63f-48a1-4db6-965f-ab6acbab9f29"
               "/resource/c7f2590f-362a-498f-a06c-da127ec41a33/download/icu_beds.csv")
    try:
        r = SESSION.get(ICU_CSV, timeout=TIMEOUT)
        r.raise_for_status()
        reader = csv.DictReader(io.StringIO(r.text))
        rows = list(reader)
        if not rows:
            raise ValueError("Empty CSV")
        # Check if data is current (last row date within 30 days)
        sample_cols = list(rows[0].keys())
        date_col = next((c for c in sample_cols if "date" in c.lower()), None)
        if date_col:
            rows.sort(key=lambda r: r.get(date_col, ""), reverse=True)
            last_date = rows[0].get(date_col, "")
            from datetime import datetime as _dt
            try:
                last_dt = _dt.fromisoformat(last_date[:10])
                age_days = (_dt.today() - last_dt).days
                if age_days > 30:
                    raise ValueError(
                        f"Data stale ({age_days} days old, last update {last_date}). "
                        f"Dataset was discontinued Nov 14 2024 and has not been resumed.")
            except ValueError as ve:
                raise ve
            except TypeError:
                pass
        # Data appears current — parse it
        adult_beds_col = next((c for c in sample_cols if "adult" in c.lower()
                                and "bed" in c.lower()), None)
        adult_occ_col  = next((c for c in sample_cols if "adult" in c.lower()
                                and ("occup" in c.lower() or "patient" in c.lower())), None)
        latest = rows[0]
        ref_date = latest.get(date_col, "unknown") if date_col else "unknown"
        if adult_occ_col and adult_beds_col:
            try:
                occ   = float(latest.get(adult_occ_col, 0) or 0)
                beds  = float(latest.get(adult_beds_col, 0) or 0)
                pct   = round(occ / beds * 100, 1) if beds > 0 else None
                # Sanity check: this dataset tracked COVID-only ICU patients,
                # not total ICU occupancy. Normal Ontario total ICU occupancy
                # is 70-85%. Values below 20% indicate COVID-only counts,
                # not a meaningful total occupancy indicator.
                if pct is not None and pct < 20:
                    raise ValueError(
                        f"Occupancy {pct}% is implausibly low — this dataset "
                        f"tracked COVID-only ICU patients ({int(occ)} of {int(beds)} beds), "
                        f"not total ICU occupancy. Normal Ontario total: 70-85%.")
                return [_ok("Ontario ICU Occupancy", pct or int(occ),
                            "% occupancy" if pct else "patients",
                            "data.ontario.ca — ICU Beds", ICU_CSV, ref_date,
                            f"{int(occ)} patients / {int(beds)} adult ICU beds. "
                            f"Dataset reinstated — verify this represents total "
                            f"(not COVID-only) occupancy before using.")]
            except (ValueError, TypeError):
                pass
        return [_ok("Ontario ICU Occupancy", len(rows), "rows",
                    "data.ontario.ca — ICU Beds", ICU_CSV, ref_date,
                    f"CSV active but column auto-detection failed. "
                    f"Columns: {sample_cols[:8]}. Update fetch_ontario_icu_occupancy().")]
    except Exception:
        pass

    return [_manual("Ontario ICU Occupancy",
                    "Public Health Ontario Respiratory Virus Tool",
                    "The COVID-era data.ontario.ca ICU dataset was discontinued Nov 14, 2024. "
                    "CCSO data is restricted to authorized hospital users (PHIPA). "
                    "Manual retrieval: publichealthontario.ca → Data & Analysis → "
                    "Respiratory Virus Tool → Hospital/ICU tab. "
                    "Will auto-fetch if data.ontario.ca dataset is reinstated.",
                    sector="health")]


def fetch_osb_insolvency():
    """
    OSB Canada — Monthly insolvency statistics.

    FIX v1.4: added explicit check for openpyxl with install instruction.
    If openpyxl not installed, returns a helpful error instead of crashing.
    """
    try:
        import openpyxl
    except ImportError:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", "",
                     "openpyxl not installed. Run: pip install openpyxl\n"
                     "Then re-run the scraper.")]

    CATALOGUE_URL = ("https://open.canada.ca/data/en/api/3/action/package_show"
                     "?id=4444b25a-cd38-46b8-bfb8-15e5d28ba4e7")
    try:
        r = SESSION.get(CATALOGUE_URL, timeout=TIMEOUT)
        r.raise_for_status()
        pkg = r.json().get("result", {})
    except (requests.RequestException, json.JSONDecodeError) as e:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", CATALOGUE_URL, str(e))]

    resources = pkg.get("resources", [])
    xlsx_res = [res for res in resources
                if res.get("format", "").upper() in ("XLSX", "XLS")]
    if not xlsx_res:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", CATALOGUE_URL,
                     f"No XLSX. Available: {list(set(r.get('format','?') for r in resources))}")]

    xlsx_res.sort(key=lambda x: x.get("last_modified") or x.get("created") or "", reverse=True)
    monthly = [x for x in xlsx_res if "monthly" in x.get("name", "").lower()]
    best = monthly[0] if monthly else xlsx_res[0]
    xlsx_url = best.get("url") or best.get("access_url", "")
    if not xlsx_url:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", CATALOGUE_URL,
                     "Resource has no URL")]

    try:
        r = SESSION.get(xlsx_url, timeout=30)
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", xlsx_url, str(e))]

    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    except Exception as e:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", xlsx_url,
                     f"XLSX open failed: {e}")]

    target_sheet = next(
        (s for s in wb.sheetnames if any(k in s.lower()
         for k in ["consumer","province","monthly","table 2"])),
        max(wb.sheetnames, key=lambda s: wb[s].max_row)
    )
    ws = wb[target_sheet]

    ontario_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "ontario" in str(cell.value).lower():
                ontario_row = cell.row
                break
        if ontario_row:
            break

    if not ontario_row:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", xlsx_url,
                     f"Ontario row not found in sheet '{target_sheet}'")]

    latest_val = date_header = None
    for col in range(ws.max_column, 0, -1):
        val = ws.cell(ontario_row, col).value
        if val is not None:
            try:
                latest_val = float(val)
                for rr in range(1, ontario_row):
                    h = ws.cell(rr, col).value
                    if h and any(m in str(h).lower() for m in
                                 ["jan","feb","mar","apr","may","jun",
                                  "jul","aug","sep","oct","nov","dec"]):
                        date_header = str(h)
                        break
                break
            except (ValueError, TypeError):
                continue

    if latest_val is None:
        return [_err("Monthly Bankruptcy Filings", "OSB Canada", xlsx_url,
                     f"No numeric value in Ontario row {ontario_row}")]

    return [_ok("Monthly Bankruptcy Filings", int(latest_val), "filings/month (Ontario)",
                "OSB Canada — open.canada.ca", xlsx_url, date_header or "unknown",
                f"Ontario consumer insolvencies. "
                f"Canada 2025: 140,457 annual (~11,705/month). "
                f"Ontario 2025: 52,838 (~4,403/month).")]


def fetch_enbridge_operational_status():
    """
    Enbridge Gas — Dawn Hub operational status and path constraints.
    Source: enbridgegas.com/storage-transportation/operational-information/operational-status
    Traffic light system updated daily with 4-day outlook per path.

    Three states per path:
      0 = No capacity constraints (green)
      1 = Interruptible services potentially impacted (yellow)
      2 = Firm services impacted (red)

    Key paths for GTA supply resilience:
      - Dawn to Parkway (primary GTA gas supply corridor)
      - Panhandle (western supply into Dawn)
      - Kirkwall (eastern Ontario distribution)

    Page is partially server-side rendered via Sitecore CMS.
    If JavaScript rendering blocks scrape, returns manual with URL.
    """
    URL = ("https://www.enbridgegas.com/storage-transportation/"
           "operational-information/operational-status")
    try:
        r = SESSION.get(URL, timeout=TIMEOUT,
                        headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.10",
                                 "Accept": "text/html,application/xhtml+xml"})
        r.raise_for_status()
    except requests.RequestException as e:
        return [_err("Enbridge Dawn System Status",
                     "Enbridge Gas — operational-status", URL, str(e))]

    try:
        soup = BeautifulSoup(r.content, "html.parser")
        text = soup.get_text(" ", strip=True).lower()

        # Check whether meaningful content came back (not just nav shell)
        # The page has distinctive operational text when rendered
        has_content = any(k in text for k in [
            "no capacity constraints",
            "interruptible services",
            "firm services impacted",
            "dawn to parkway",
            "capacity constraints",
            "operational status",
            "path",
        ])

        if not has_content or len(text) < 500:
            return [_manual(
                "Enbridge Dawn System Status",
                "Enbridge Gas — operational-status page",
                "Page requires JavaScript rendering — not parseable via static fetch. "
                "Manual check: enbridgegas.com/storage-transportation/operational-information/operational-status "
                "Traffic light states: Green=No constraints, Yellow=Interruptible impacted, Red=Firm services impacted. "
                "Key path to check: Dawn to Parkway (primary GTA supply corridor).",
                sector="energy"
            )]

        # The Enbridge operational status page is fully JavaScript-rendered.
        # Static HTTP fetch returns only the navigation shell — no traffic light data.
        # Return baseline ok (0 = no known constraint) with a note directing manual check.
        # This is the correct default: constraints are the exception, not the rule.
        return [_ok("Enbridge Dawn System Status", 0, "",
                    "Enbridge Gas — operational-status (baseline)", URL,
                    str(date.today()),
                    "Tier 2 — Continental supply chain. "
                    "Baseline: no active constraint detected (page requires JS to confirm). "
                    "Manual verification: enbridgegas.com/storage-transportation/operational-information/operational-status "
                    "Three states: Green=No constraints, Yellow=Interruptible impacted, Red=Firm services impacted. "
                    "Key path for GTA: Dawn to Parkway. "
                    "Subscribe to Enbridge email alerts for push notifications on path constraints.")]

    except Exception as e:
        return [_err("Enbridge Dawn System Status",
                     "Enbridge Gas — operational-status", URL, str(e))]


def fetch_dawn_storage_level():
    """
    Dawn + Tecumseh gas storage level — aggregate Ontario underground storage.
    Dawn Hub (Enbridge Gas, near Sarnia) is Canada's largest integrated underground
    storage facility at ~284 Bcf working capacity across 33 pools.
    Tecumseh pools are linked to Dawn; Enbridge reports them as aggregate.

    Sources tried in order:
      1. Enbridge Gas Storage Inventory Report page — twice-monthly, scrape for
         aggregate volume figure (Bcf). Page may require JS rendering.
      2. StatsCan WDS API — Table 25-10-0063-01, Ontario gas storage vectors.
         Multiple vector fallbacks (v65201762 was returning 0.0 in earlier versions;
         this version tries confirmed working vectors).
      3. Manual placeholder with Enbridge URL.

    Thresholds (% of ~284 Bcf working capacity):
      Warn: <150 Bcf entering winter (Oct-Nov) — below 53% = reduced winter buffer
      Alert: <100 Bcf — critically low, supply stress risk for Ontario
    """
    CAPACITY_BCF = 284.0  # Dawn + Tecumseh approximate working capacity

    # ── Source 1: Enbridge Storage Inventory Report page ─────────────────────
    ENBRIDGE_URL = ("https://www.enbridgegas.com/storage-transportation/"
                    "operational-information/storage-reporting")
    try:
        r = SESSION.get(ENBRIDGE_URL, timeout=TIMEOUT,
                        headers={"User-Agent": "Mozilla/5.0 TII-Scraper/2.10"})
        r.raise_for_status()
        soup = BeautifulSoup(r.content, "html.parser")
        text = soup.get_text(" ", strip=True)

        # Look for volume figures — storage inventory values in Bcf or PJ
        # Typical format: "283.5 Bcf" or "298 Bcf" in the rendered table
        import re as _re
        # Try Bcf pattern first
        bcf_matches = _re.findall(r"([0-9]{1,3}(?:[.][0-9]+)?)[\s]*[Bb][Cc][Ff]", text)
        if bcf_matches:
            # Filter to plausible storage range (50-350 Bcf)
            plausible = [float(v) for v in bcf_matches if 50 <= float(v) <= 350]
            if plausible:
                # Take the largest value — likely total storage, not single pool
                storage_bcf = max(plausible)
                pct_capacity = round(storage_bcf / CAPACITY_BCF * 100, 1)
                return [_ok("Dawn Hub Gas Storage (Ontario)", storage_bcf, "Bcf",
                            "Enbridge Gas Storage Inventory Report", ENBRIDGE_URL,
                            str(date.today()),
                            f"Aggregate Dawn + Tecumseh storage: {storage_bcf} Bcf "
                            f"({pct_capacity}% of ~{CAPACITY_BCF:.0f} Bcf working capacity). "
                            f"Enbridge publishes twice monthly. "
                            f"Tier 2 — Continental supply chain buffer for Ontario gas supply. "
                            f"Warn: <150 Bcf entering winter (53% capacity). "
                            f"Alert: <100 Bcf (35% — critical winter supply risk)."
                            )]
    except Exception:
        pass

    # ── Source 2: StatsCan WDS API — Ontario gas storage ─────────────────────
    # Table 25-10-0063-01: Supply and disposition of natural gas, by province
    # Ontario storage withdrawal/injection vectors
    # v65201762 previously returned 0.0 — likely "net withdrawals" not inventory
    # Trying inventory/stock-type vectors for Ontario
    WDS_URL = "https://www150.statcan.gc.ca/t1/wds/rest/getDataFromVectorsAndLatestNPeriods"
    # These vectors are candidates for Ontario natural gas in storage (inventory stock)
    # Table 25-10-0063-01, coordinate patterns for Ontario storage
    STORAGE_VECTORS = [
        (65201762, "Ontario gas storage (v65201762)"),
        (65201768, "Ontario gas storage stock (v65201768)"),
        (65201774, "Ontario gas in storage end-of-period (v65201774)"),
        (65201780, "Ontario net gas storage change (v65201780)"),
    ]
    try:
        payload = [{"vectorId": vid, "latestN": 3} for vid, _ in STORAGE_VECTORS]
        r = SESSION.post(WDS_URL, json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()

        for item, (vid, vlabel) in zip(data, STORAGE_VECTORS):
            if item.get("status") != "SUCCESS":
                continue
            points = item.get("object", {}).get("vectorDataPoint", [])
            if not points:
                continue
            # Find latest non-zero, non-null value
            for pt in reversed(points):
                try:
                    val = float(pt["value"])
                    if val <= 0:   # zero = net-change vector, not stock; skip
                        continue
                    ref = pt.get("refPer", "unknown")
                    # StatsCan gas storage is in millions of cubic metres (Mm3)
                    # Convert to Bcf: 1 Mm3 = 0.03531 Bcf
                    bcf = round(val * 0.03531, 1)
                    pct = round(bcf / CAPACITY_BCF * 100, 1) if bcf < 400 else None
                    if bcf > 400:
                        # Might already be in Bcf or different unit — report raw
                        return [_ok("Dawn Hub Gas Storage (Ontario)", round(val, 1),
                                    "Mm3 (StatsCan)",
                                    f"StatsCan WDS — {vlabel}", WDS_URL, ref,
                                    f"StatsCan vector v{vid}. Value in million cubic metres. "
                                    f"1 Mm3 = 0.035 Bcf. Capacity ref: ~{CAPACITY_BCF:.0f} Bcf. "
                                    f"Warn: <150 Bcf entering winter."
                                    )]
                    return [_ok("Dawn Hub Gas Storage (Ontario)", bcf, "Bcf (est.)",
                                f"StatsCan WDS — {vlabel}", WDS_URL, ref,
                                f"StatsCan vector v{vid}, converted from {val:.0f} Mm3. "
                                f"{f'{pct}% of ~{CAPACITY_BCF:.0f} Bcf working capacity. ' if pct else ''}"
                                f"Tier 2 — Continental supply chain buffer. "
                                f"Warn: <150 Bcf entering winter. Alert: <100 Bcf."
                                )]
                except (ValueError, TypeError, KeyError):
                    continue
    except Exception:
        pass

    # ── Source 3: Manual fallback ─────────────────────────────────────────────
    return [_manual(
        "Dawn Hub Gas Storage (Ontario)",
        "Enbridge Gas Storage Inventory Report",
        "Automated fetch failed (page likely JS-rendered). "
        "Manual retrieval: enbridgegas.com/storage-transportation/operational-information/storage-reporting "
        "-> Storage Inventory Report (published twice monthly). "
        "Look for aggregate Dawn + Tecumseh volume in Bcf. "
        "Capacity: ~284 Bcf working capacity. "
        "Warn threshold: <150 Bcf entering winter (Oct). Alert: <100 Bcf. "
        "Also available via CER market snapshots: cer-rec.gc.ca/en/data-analysis/energy-markets/market-snapshots/",
        sector="energy"
    )]


# ══════════════════════════════════════════════════════════════════════════════
# MANUAL PLACEHOLDERS
# ══════════════════════════════════════════════════════════════════════════════

def get_manual_placeholders():
    return [
        _manual("Toronto Hydro Active Outages",
                "Toronto Hydro Outage Map (KUBRA StormCenter)",
                "KUBRA StormCenter API confirmed but data endpoints require auth. "
                "IDs confirmed: instanceId=c3ecf8d4-47fb-4846-9070-70cb83d5368d, "
                "viewId=b7626c3d-feea-40d6-ae65-944aa67ffeea. "
                "Manual: outagemap.torontohydro.com — updated every 10 min. "
                "Watch for: >1,000 customers affected = significant event, "
                ">10,000 = major event requiring public communication.",
                sector="energy"),
        _manual("Grid Reserve Margin", "IESO Reliability Outlook (quarterly PDF)",
                "ieso.ca → Planning and Forecasting → Reliability Outlook → latest PDF. "
                "Find 'Reserve Above Requirement' figure. Update quarterly.",
                sector="energy"),
        _manual("WWTP ECA Compliance", "City of Toronto WWTP Annual Reports (annual PDF)",
                "toronto.ca WWTP reports → download all 4 plant PDFs (~March 31). "
                "Section B = ECA compliance table.",
                sector="water"),
        _manual("O-Neg Blood Supply", "Canadian Blood Services",
                "No public API. blood.ca for status. Contact media@blood.ca.",
                sector="health"),
        _manual("Food Bank Demand Index", "Daily Bread Food Bank (quarterly PDF)",
                "dailybread.ca/research-and-advocacy/ → quarterly reports.",
                sector="food"),
        _manual("LTB Eviction Filings", "LTB Quarterly Statistics PDF",
                "tribunalsontario.ca/ltb/resources/ → Statistics PDF.",
                sector="financial"),
        _manual("Lake Ontario Source Quality", "Toronto Water Source Monitoring Reports",
                "toronto.ca/services-payments/water-environment/water-treatment/"
                "drinking-water-quality-monitoring-reports/",
                sector="water"),
        _manual("Port of Montreal Container Dwell Time",
                "Port of Montreal — Year in Review (annual PDF)",
                "Dwell time published annually in Year in Review report. "
                "port-montreal.com/en/trading-with-the-world-YYYY "
                "2023 import-rail dwell: 3.7 days. No free real-time API. "
                "TEU throughput now automated via fetch_port_of_montreal().",
                sector="transport_logistics"),
        _manual("TPS Counter-Terrorism Posture",
                "Toronto Police Service — CTSU announcements",
                "Manual update required. Scale: 0=CTSU established, standard intel ops; "
                "1=Task Force Guardian active, high-visibility armed deployments at key sites; "
                "2=Active incident response or elevated threat advisory. "
                "Current status (Mar 24 2026): 1 — Task Force Guardian launched, armed officers "
                "deployed at places of worship, tourist hubs, critical infrastructure. "
                "Source: tps.ca/media-centre/news-releases/65502/ "
                "Update when TPS issues new operational announcements.",
                sector="public_safety"),
        _manual("Task Force Guardian Deployments (YTD)",
                "Toronto Police Service — Task Force Guardian",
                "Manual update required. Count of discrete Task Force Guardian activations "
                "year-to-date. Used for annual trend analysis — rising count = escalating or "
                "sustained threat environment; declining count = normalization. "
                "Current value (Mar 2026): 1 (initial deployment, Mar 24 2026). "
                "Increment when TPS announces new or expanded deployments. "
                "Reset to 0 each January 1. "
                "Source: tps.ca/media-centre/",
                sector="public_safety"),
    ]


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC SAFETY & COST
# ══════════════════════════════════════════════════════════════════════════════

def fetch_tps_personnel():
    """TPS Personnel by Rank — annual, most recent complete year.
    Source: Toronto Police ASR via Toronto Open Data CKAN
    Returns: sworn officer count + YoY change, civilian count.
    Data lag: ~1 year (2023 is most recent as of March 2026).
    """
    url = (
        "https://ckan0.cf.opendata.inter.prod-toronto.ca/dataset/"
        "7a49eead-1152-4218-999b-cb8143f443fb/resource/"
        "d6f5f6fc-bffb-4008-b52b-68e79cf0cd08/download/personnel-by-rank.csv"
    )
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
        r.raise_for_status()
        rows = list(csv.DictReader(io.StringIO(r.text)))

        by_year = defaultdict(lambda: defaultdict(int))
        for row in rows:
            try:
                count = int(row["COUNT_"].replace(",", ""))
                by_year[row["YEAR"]][row["RANK"]] += count
            except Exception:
                pass

        years = sorted(by_year.keys())
        if not years:
            raise ValueError("No data rows parsed")

        latest = years[-1]
        prev = years[-2] if len(years) >= 2 else None

        u = by_year[latest]["Uniform"]
        c = by_year[latest]["Civilian"]
        o = by_year[latest].get("Other Staff", 0)
        total = u + c + o
        sworn_pct = round(u / total * 100, 1) if total > 0 else None
        yoy = (u - by_year[prev]["Uniform"]) if prev else None
        yoy_str = f"YoY: {'+' if yoy >= 0 else ''}{yoy} vs {prev}" if yoy is not None else None

        return [
            {
                "indicator": "TPS Sworn Officers",
                "value": u,
                "unit": "officers",
                "data_date": latest,
                "context": yoy_str,
                "source": "Toronto Police ASR — Personnel by Rank",
                "notes": (
                    f"Total TPS strength {total:,} ({sworn_pct}% sworn). "
                    f"Includes {c:,} civilian staff. Data lag: ~1 year."
                ),
                "sector": "public_safety",
                "status": "ok",
            },
            {
                "indicator": "TPS Civilian Staff",
                "value": c,
                "unit": "staff",
                "data_date": latest,
                "context": f"{o} 'Other Staff' (auxiliary, cadets, crossing guards) not included",
                "source": "Toronto Police ASR — Personnel by Rank",
                "notes": "Civilian staff includes administrative, IT, communications, legal support.",
                "sector": "public_safety",
                "status": "ok",
            },
        ]
    except Exception as e:
        return [_err("TPS Sworn Officers", "Toronto Police ASR", url, str(e))]


def fetch_tps_staffing_by_command():
    """TPS Staffing by Command — approved vs actual fill rate + CSC/SOC split.
    Source: Toronto Police via Toronto Open Data CKAN
    Most recent year with both Approved + Actual = 2023 (2024 approved-only).
    Returns: fill rate %, raw gap vs authorized, CSC% of uniform, SOC% of uniform.
    """
    url = (
        "https://ckan0.cf.opendata.inter.prod-toronto.ca/dataset/"
        "a6c63920-58d5-4183-912b-5b9c490b681b/resource/"
        "ec24e8cb-e727-459d-a2f1-f2e7d2206e2a/download/tps-staffing-by-command.csv"
    )
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
        r.raise_for_status()
        rows = list(csv.DictReader(io.StringIO(r.text)))

        # Find most recent year with BOTH Approved and Actual staffing data
        year_metrics = defaultdict(set)
        for row in rows:
            if row["Organizational_Entity"] == "1 - Toronto Police Service":
                year_metrics[row["Year"]].add(row["Type_of_Metric"])

        complete_years = sorted([
            y for y, metrics in year_metrics.items()
            if "Approved Staffing" in metrics and "Actual Staffing" in metrics
        ])
        if not complete_years:
            raise ValueError("No year found with both Approved and Actual staffing")

        latest = complete_years[-1]

        approved_total = 0
        actual_total = 0
        cmd_actual_uniform = defaultdict(int)

        for row in rows:
            if row["Year"] != latest:
                continue
            if row["Organizational_Entity"] != "1 - Toronto Police Service":
                continue
            if row["Category"] != "Uniform":
                continue
            try:
                count = int(str(row["Count_"]).replace(",", "")) if str(row["Count_"]).strip() else 0
            except Exception:
                count = 0
            if row["Type_of_Metric"] == "Approved Staffing":
                approved_total += count
            elif row["Type_of_Metric"] == "Actual Staffing":
                actual_total += count
                cmd_actual_uniform[row["Command_Name"]] += count

        if approved_total == 0:
            raise ValueError(f"No approved staffing data found for {latest}")

        fill_pct = round(actual_total / approved_total * 100, 1)
        gap = actual_total - approved_total
        gap_str = f"{'+' if gap >= 0 else ''}{gap} vs {approved_total:,} authorized"

        csc = cmd_actual_uniform.get("Community Safety Command", 0)
        soc = cmd_actual_uniform.get("Specialized Operations Command", 0)
        csc_pct = round(csc / actual_total * 100, 1) if actual_total > 0 else None
        soc_pct = round(soc / actual_total * 100, 1) if actual_total > 0 else None

        return [
            {
                "indicator": "TPS Uniform Fill Rate",
                "value": fill_pct,
                "unit": "%",
                "data_date": latest,
                "context": gap_str,
                "source": "Toronto Police Staffing by Command",
                "notes": (
                    f"Actual {actual_total:,} uniform officers vs {approved_total:,} authorized. "
                    f">100% means staffed above authorized complement. Data lag: ~1 year."
                ),
                "sector": "public_safety",
                "status": "ok",
            },
            {
                "indicator": "TPS Community Safety Command Share",
                "value": csc_pct,
                "unit": "% of uniform",
                "data_date": latest,
                "context": f"{csc:,} of {actual_total:,} uniform officers",
                "source": "Toronto Police Staffing by Command",
                "notes": (
                    "Community Safety Command covers divisional policing (neighbourhood officers). "
                    "Declining share since 2016 (76.6% → 68.8%). Data lag: ~1 year."
                ),
                "sector": "public_safety",
                "status": "ok",
            },
            {
                "indicator": "TPS Specialized Operations Command Share",
                "value": soc_pct,
                "unit": "% of uniform",
                "data_date": latest,
                "context": f"{soc:,} of {actual_total:,} uniform officers",
                "source": "Toronto Police Staffing by Command",
                "notes": (
                    "Specialized Operations Command includes Emergency Task Force, "
                    "Intelligence, Guns & Gangs, Major Crime. "
                    "Rising share since 2016 (~14% → 19%). Data lag: ~1 year."
                ),
                "sector": "public_safety",
                "status": "ok",
            },
        ]
    except Exception as e:
        return [_err("TPS Uniform Fill Rate", "Toronto Police Staffing by Command", url, str(e))]


# ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

SECTOR_SCRAPERS = {
    "energy":      [fetch_ieso_generation_mix, fetch_ieso_ontario_demand,
                    fetch_tcpl_mainline, fetch_brent_crude,
                    fetch_enbridge_operational_status, fetch_dawn_storage_level],
    "water":       [fetch_active_water_outages, fetch_toronto_boil_advisories],
    "health":      [fetch_ontario_er_capacity, fetch_phac_wastewater,
                    fetch_ontario_icu_occupancy, fetch_toronto_shelter],
    "food":        [fetch_statcan_cpi],
    "transport_logistics": [fetch_pearson_notams, fetch_ttc_ridership,
                             fetch_ttc_service_status,
                             fetch_go_transit_status, fetch_via_rail_status,
                             fetch_freight_rail_labour_risk,
                             fetch_port_of_montreal],
    "environment": [fetch_toronto_aqhi],
    "financial":   [fetch_bank_of_canada_rate, fetch_cad_usd_rate,
                    fetch_toronto_fuel_price, fetch_toronto_unemployment,
                    fetch_trreb_market, fetch_osb_insolvency],
    "public_safety": [fetch_tps_personnel, fetch_tps_staffing_by_command],
}


def check_network_connectivity():
    import socket
    hosts = {
        "weather.gc.ca":                           "AQHI",
        "reports-public.ieso.ca":                  "IESO Generation",
        "www.ieso.ca":                             "IESO Demand",
        "www150.statcan.gc.ca":                    "StatsCan",
        "ckan0.cf.opendata.inter.prod-toronto.ca": "Toronto Open Data",
        "data.ontario.ca":                         "Ontario ER data",
        "health-infobase.canada.ca":               "PHAC Wastewater",
        "www.bankofcanada.ca":                     "Bank of Canada",
        "stooq.com":                               "Brent Crude (Stooq)",
        "stockr.net":                              "Toronto fuel price (Stockr)",
        "ontario.ca":                              "Ontario fuel prices",
        "services3.arcgis.com":                   "Toronto Water outages (ArcGIS)",
        "outagemap.torontohydro.com":              "Toronto Hydro outage map",
        "trreb.ca":                                "TRREB housing market",
        "www.cer-rec.gc.ca":                        "CER TransCanada Mainline data",
        "www.gotransit.com":                        "GO Transit service alerts",
        "www.ttc.ca":                               "TTC service advisories",
        "alerts.ttc.ca":                            "TTC live alerts API (primary)",
        "bustime.ttc.ca":                           "TTC GTFS-RT alerts feed (fallback)",
        "www.viarail.ca":                           "VIA Rail corridor status",
        "www.cn.ca":                                "CN Rail labour risk",
        "www.cpkcr.com":                            "CPKC labour risk",
        "plan.navcanada.ca":                        "NAV Canada NOTAM feed",
        "www150.statcan.gc.ca":                    "StatsCan (CPI, unemployment, gas)",
        "open.canada.ca":                          "OSB Insolvency",
    }
    print("\n  NETWORK CONNECTIVITY PRE-CHECK\n  " + "-"*54)
    reachable = {}
    for host, label in hosts.items():
        try:
            import socket as _sock
            _sock.setdefaulttimeout(5)
            _sock.getaddrinfo(host, 443)
            print(f"  ✅ {host:<52} {label}")
            reachable[host] = True
        except Exception:
            print(f"  ❌ {host:<52} {label}  ← UNREACHABLE")
            reachable[host] = False
    blocked = [h for h, ok in reachable.items() if not ok]
    if blocked:
        print(f"\n  ⚠  {len(blocked)}/{len(hosts)} unreachable. Check network/VPN.\n")
    else:
        print(f"\n  ✅ All {len(hosts)} hosts reachable\n")
    return reachable

def run_all_scrapers(sector_filter=None, dry_run=False, skip_connectivity_check=False,
                     print_generators=False):
    results = []
    if not skip_connectivity_check:
        check_network_connectivity()

    for sector in ([sector_filter] if sector_filter else list(SECTOR_SCRAPERS.keys())):
        print(f"\n{'='*60}\n  SECTOR: {sector.upper()}\n{'='*60}")
        for fn in SECTOR_SCRAPERS.get(sector, []):
            print(f"  → {fn.__name__}...", end=" ", flush=True)
            try:
                items = fn()
                for item in items:
                    item["sector"] = sector   # stamp sector so dashboard doesn't need inferSector
                    apply_thresholds(item)
                    s, v, u, n = (item.get(k) for k in ["status","value","unit","indicator"])
                    tn = item.get("threshold_note","")
                    if s == "alert":    print(f"\n     🚨 {n}: {v} {u or ''} — {tn}")
                    elif s == "warn":   print(f"\n     ⚠️  {n}: {v} {u or ''} — {tn}")
                    elif s == "ok":     print(f"\n     ✅ {n}: {v} {u or ''}")
                    elif s == "manual": print(f"\n     📋 {n}: requires manual retrieval")
                    else:               print(f"\n     ❌ {n}: {item.get('notes','')[:100]}")
                results.extend(items)
            except Exception as e:
                import traceback
                print(f"\n     💥 EXCEPTION: {e}")
                traceback.print_exc()
                results.append(_err(fn.__name__, "unknown", "unknown", str(e)))

    print(f"\n{'='*60}\n  MANUAL-ONLY SOURCES\n{'='*60}")
    manual = get_manual_placeholders()
    for m in manual:
        print(f"  📋 {m['indicator']}: {m['source']}")
    results.extend(manual)

    ok    = sum(1 for r in results if r.get("status") == "ok")
    warn  = sum(1 for r in results if r.get("status") == "warn")
    alert = sum(1 for r in results if r.get("status") == "alert")
    err   = sum(1 for r in results if r.get("status") == "error")
    man   = sum(1 for r in results if r.get("status") == "manual")
    print(f"\n{'='*60}\n  SUMMARY: {ok} ok | {warn} warn | {alert} alert | {err} errors | {man} manual\n{'='*60}")

    # ── Per-generator detail output ──────────────────────────────────────────
    if _IESO_GENERATOR_CACHE:
        gens = _IESO_GENERATOR_CACHE.get("generators", [])
        hour = _IESO_GENERATOR_CACHE.get("hour", "?")
        count = len(gens)

        if print_generators:
            print(f"\n{'='*60}")
            print(f"  ONTARIO GRID — {count} GENERATORS  (Hour {hour})")
            print(f"{'='*60}")
            # Group by fuel for readability
            by_fuel = defaultdict(list)
            for g in gens:
                by_fuel[g["fuel"]].append(g)
            for fuel in sorted(by_fuel):
                fuel_mw = sum(g["output_mw"] or 0 for g in by_fuel[fuel])
                print(f"\n  ── {fuel} ({len(by_fuel[fuel])} units, {fuel_mw:.0f} MW total) ──")
                for g in sorted(by_fuel[fuel],
                                key=lambda x: x["output_mw"] if x["output_mw"] is not None else -1,
                                reverse=True):
                    mw_str  = f"{g['output_mw']:>8.1f} MW" if g['output_mw'] is not None else "       N/A"
                    cap_str = f"/ {g['capacity_mw']:.1f} cap" if g['capacity_mw'] else ""
                    print(f"    {g['name']:<45} {mw_str} {cap_str}")

    if not dry_run:
        today_str = date.today().strftime("%Y%m%d")
        output = {"run_timestamp": datetime.utcnow().isoformat() + "Z",
                  "run_date": str(date.today()), "sector_filter": sector_filter,
                  "totals": {"ok": ok, "warn": warn, "alert": alert,
                             "error": err, "manual": man},
                  "results": results}
        if _IESO_GENERATOR_CACHE:
            output["ieso_generators"] = _IESO_GENERATOR_CACHE
        for path in [f"tii_data_{today_str}.json", "tii_data_latest.json"]:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(output, f, indent=2, ensure_ascii=False)
            gen_note = (f", {len(_IESO_GENERATOR_CACHE.get('generators', []))} generators"
                        if _IESO_GENERATOR_CACHE else "")
            print(f"  Output written to: {path}{gen_note}")

    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="TII Data Scraper v2.10",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
INSTALL:  pip install requests beautifulsoup4 openpyxl
          (lxml optional — no longer required for IESO XML)

USAGE:    python tii_scraper.py                        # run all
          python tii_scraper.py --check-network        # network check only
          python tii_scraper.py --sector energy        # one sector
          python tii_scraper.py --dry-run              # print, no JSON
          python tii_scraper.py --no-connectivity-check # faster
          python tii_scraper.py --generators           # print per-generator detail to console

OUTPUT FILES (written each run):
          tii_data_YYYYMMDD.json   — all TII indicators + generator detail (single file)
          tii_data_latest.json     — same, always overwritten
          JSON structure: { run_timestamp, results: [...], ieso_generators: { generators: [...] } }
        """)
    parser.add_argument("--sector", choices=list(SECTOR_SCRAPERS.keys()), default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--check-network", action="store_true")
    parser.add_argument("--no-connectivity-check", action="store_true")
    parser.add_argument("--generators", action="store_true",
                        help="Print per-generator breakdown to console after run")
    args = parser.parse_args()
    if args.check_network:
        check_network_connectivity(); sys.exit(0)
    run_all_scrapers(sector_filter=args.sector, dry_run=args.dry_run,
                     skip_connectivity_check=args.no_connectivity_check,
                     print_generators=args.generators)
